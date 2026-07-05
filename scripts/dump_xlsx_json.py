#!/usr/bin/env python3
"""Serializa contenido de xlsx en carpeta a ``xlsx_dump.json`` para inspección.

Incluye ``Plantilla*.xlsx`` de la raíz; captura errores de lectura por archivo.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from e2e_common import add_folder_argument, resolve_e2e_folder

OUT = ROOT / "xlsx_dump.json"


def dump_xlsx(path: Path) -> dict:
    """Serializa contenido de un xlsx a estructura dict para inspección JSON."""
    result: dict = {"name": path.name, "path": str(path)}
    try:
        df = pd.read_excel(path, header=0)
        result["extract_rows"] = df.fillna("").astype(str).to_dict(orient="records")
        result["extract_columns"] = list(df.columns)
    except Exception as exc:
        result["extract_error"] = str(exc)

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        result["sheet"] = ws.title
        result["header"] = {}
        for addr in ["B7", "B9", "B11", "B13"]:
            result["header"][addr] = ws[addr].value
        rows = []
        for row_idx in range(1, ws.max_row + 1):
            vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 13)]
            if any(v is not None for v in vals):
                rows.append({"row": row_idx, "values": vals})
        result["rows"] = rows
        wb.close()
    except Exception as exc:
        result["workbook_error"] = str(exc)
    return result


def main() -> int:
    """Punto de entrada CLI del script."""
    parser = argparse.ArgumentParser()
    add_folder_argument(parser)
    args = parser.parse_args()

    data: dict = {"files": []}
    folder = resolve_e2e_folder(args.folder)
    data["folder"] = folder.name
    for x in sorted(folder.glob("*.xlsx")):
        data["files"].append(dump_xlsx(x))
    plantilla = next(ROOT.glob("Plantilla*.xlsx"), None)
    if plantilla:
        data["plantilla"] = dump_xlsx(plantilla)
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
