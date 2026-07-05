#!/usr/bin/env python3
"""Inspecciona archivos xlsx: hojas, columnas y fórmulas de footer (primeras 25 filas).

Herramienta de debug; default ``tests/fixtures/demo_card``.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))


def inspect_workbook(path: Path) -> None:
    """Imprime hojas, encabezados y fórmulas de un libro Excel."""
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        print(f"Hoja: {ws.title} ({ws.max_row} filas x {ws.max_column} cols)")
        headers = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column, 6) + 1)]
        print(f"  Fila 1: {headers}")
        for row in range(1, min(ws.max_row, 25) + 1):
            formulas = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{cell.coordinate}={cell.value}")
            if formulas:
                print(f"  Fórmulas fila {row}: {formulas}")
    wb.close()


def main() -> int:
    """Punto de entrada CLI del script."""
    parser = argparse.ArgumentParser()
    parser.add_argument("path", type=Path, nargs="?", help="Archivo o carpeta")
    args = parser.parse_args()
    target = args.path or ROOT / "tests" / "fixtures" / "demo_card"
    if target.is_dir():
        files = sorted(target.glob("*.xlsx"))
    else:
        files = [target]
    if not files:
        print(f"No hay xlsx en {target}")
        return 1
    for file in files:
        inspect_workbook(file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
