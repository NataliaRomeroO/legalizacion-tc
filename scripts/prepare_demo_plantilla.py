#!/usr/bin/env python3
"""Genera plantilla Excel demo anonimizada para el repo a partir de una plantilla fuente.

Uso (ruta a plantilla real fuera del repo):

    python scripts/prepare_demo_plantilla.py "/ruta/Plantilla Formato....xlsx"

Salida:
- ``Plantilla Legalizacion TC Demo.xlsx`` (raíz del repo)
- ``tests/fixtures/demo_card/plantilla_base.xlsx``

Preserva fila 18 (headers), footer (Valor total / Extractos / Checkpoint) y detección
de ``template_layout``; cambia colores, título y datos identificables.
"""

from __future__ import annotations

import argparse
import glob
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from legalizacion_tc.template_layout import detect_template_layout

HEADER_ROW = 18
DATA_START_ROW = 19


def sanitize_plantilla(source: Path) -> None:
    """Anonimiza y aplica estilo demo; escribe en fixture y raíz del repo."""
    wb = load_workbook(source)
    ws = wb.active

    ws["A1"] = "Formato Demo — Legalización TC (tarjeta 1111)"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")

    ws["B7"] = None
    ws["B9"] = "Empresa Demo S.A.S"
    ws["B11"] = "Demo User A"
    ws["B13"] = "LEGALIZACIÓN TC 1111 — DEMO"

    section_fill = PatternFill("solid", fgColor="E8EEF4")
    for cell_ref in ("A5", "A17"):
        cell = ws[cell_ref]
        cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
        cell.fill = section_fill

    header_fill = PatternFill("solid", fgColor="D6E4F0")
    for col in range(1, 14):
        cell = ws.cell(row=HEADER_ROW, column=col)
        if cell.value:
            cell.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx in range(DATA_START_ROW, DATA_START_ROW + 5):
        for col in range(1, 14):
            ws.cell(row=row_idx, column=col).value = None

    replacements = {
        "Santiago Rodriguez Ruiz": "Aprobador Demo",
        "William Camilo Sanchez Osorio": "Tramitador Demo",
    }
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=14):
        for cell in row:
            if isinstance(cell.value, str):
                for old, new in replacements.items():
                    if old in cell.value:
                        cell.value = cell.value.replace(old, new)

    for row_idx in range(25, 36):
        for col_idx in range(1, 14):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and val.startswith("=+B11"):
                ws.cell(row=row_idx, column=col_idx).value = "Demo User A"

    for row_idx in range(DATA_START_ROW, 35):
        label_g = ws.cell(row=row_idx, column=7).value
        if isinstance(label_g, str) and label_g.strip().lower() == "valor total":
            for col in range(8, 13):
                ws.cell(row=row_idx, column=col).value = None
            break

    for row_idx in range(DATA_START_ROW, 35):
        for col_idx in range(1, 14):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and val.strip().lower() in (
                "extractos o movimientos",
                "checkpoint",
            ):
                for c in range(col_idx + 1, 14):
                    v2 = ws.cell(row=row_idx, column=c).value
                    if isinstance(v2, str) and v2.startswith("="):
                        ws.cell(row=row_idx, column=c).value = None

    if ws.max_row > 45:
        ws.delete_rows(46, ws.max_row - 45)

    layout = detect_template_layout(ws)
    if layout.col_usd != 8 or layout.col_cops != 10:
        raise RuntimeError(f"Layout incompatible con pipeline: {layout}")

    out_fixture = ROOT / "tests/fixtures/demo_card/plantilla_base.xlsx"
    out_root = ROOT / "Plantilla Legalizacion TC Demo.xlsx"
    out_fixture.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_fixture)
    shutil.copy2(out_fixture, out_root)
    wb.close()
    print(f"OK: {out_fixture}")
    print(f"OK: {out_root}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera plantilla demo para el repo")
    parser.add_argument(
        "source",
        nargs="?",
        help="Ruta al xlsx fuente (plantilla real). Si se omite, no regenera.",
    )
    args = parser.parse_args(argv)
    if not args.source:
        parser.error("Indique la ruta a la plantilla fuente")
    source = Path(args.source).expanduser()
    if not source.exists():
        matches = glob.glob(str(source))
        if not matches:
            print(f"No existe: {source}", file=sys.stderr)
            return 1
        source = Path(matches[0])
    sanitize_plantilla(source)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
