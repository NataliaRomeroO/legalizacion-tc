#!/usr/bin/env python3
"""Inspección E2E read-only: extracto, referencia manual y plantilla de una carpeta local.

Requiere ``Mov TC*.xlsx``; escribe log en ``e2e_run_log.txt``.
"""

from __future__ import annotations

import argparse
import sys
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

from e2e_common import add_folder_argument, infer_card_from_folder, resolve_e2e_folder

LOG = ROOT / "e2e_run_log.txt"


def log(msg: str) -> None:
    """Imprime mensaje y lo agrega al archivo de log E2E."""
    print(msg)
    with LOG.open("a", encoding="utf-8") as f:
        f.write(msg + "\n")


def main() -> int:
    """Punto de entrada CLI del script."""
    parser = argparse.ArgumentParser(description="Inspeccionar carpeta de legalización TC")
    add_folder_argument(parser)
    args = parser.parse_args()

    if LOG.exists():
        LOG.unlink()
    log(f"ROOT: {ROOT}")

    folder = resolve_e2e_folder(args.folder)
    log(f"Carpeta: {folder.name}")
    for f in sorted(folder.iterdir()):
        log(f"  - {f.name} ({f.stat().st_size} bytes)")

    import pandas as pd
    from openpyxl import load_workbook

    from legalizacion_tc.reference_loader import find_reference_legalization

    xlsx_files = sorted(folder.glob("*.xlsx"))
    log(f"\nArchivos xlsx: {[x.name for x in xlsx_files]}")

    extract = next((x for x in xlsx_files if "Mov TC" in x.name), None)
    card = infer_card_from_folder(folder) or "1111"
    reference = find_reference_legalization(folder, card)
    if extract is None:
        log("ERROR: No extracto Mov TC")
        return 1

    log(f"\n=== Extracto: {extract.name} ===")
    df = pd.read_excel(extract, header=0)
    log(f"Columnas: {list(df.columns)}")
    log(f"Filas: {len(df)}")
    log(df.to_string())

    if reference:
        log(f"\n=== Referencia: {reference.name} ===")
        wb = load_workbook(reference, data_only=True)
        ws = wb.active
        log(f"Hoja: {ws.title}, max_row={ws.max_row}")
        log(f"B7={ws['B7'].value}, B9={ws['B9'].value}, B11={ws['B11'].value}, B13={ws['B13'].value}")
        for row in range(18, min(ws.max_row + 1, 35)):
            vals = [ws.cell(row=row, column=c).value for c in range(1, 13)]
            if any(v is not None for v in vals):
                log(f"  Fila {row}: {vals}")
        wb.close()

    template_candidates = [
        folder / "plantilla_base.xlsx",
        *sorted(ROOT.glob("Plantilla*.xlsx")),
    ]
    template = next((p for p in template_candidates if p.exists()), None)
    log(f"\nPlantilla: {template}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        log(traceback.format_exc())
        raise
