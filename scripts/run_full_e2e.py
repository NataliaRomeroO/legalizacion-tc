#!/usr/bin/env python3
"""E2E completo: ejecuta pipeline local y compara Excel generado vs referencia manual.

Escribe ``e2e_results.json`` y log; exit 1 si pipeline falla o hay diffs en filas/header.
Compara hasta fila ``Valor total``; normaliza floats a 2 decimales.
"""

from __future__ import annotations

import argparse
import json
import sys
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

from e2e_common import add_folder_argument, resolve_e2e_folder

RESULT = ROOT / "e2e_results.json"
LOG: list[str] = []


def log(msg: str) -> None:
    """Imprime mensaje y lo agrega al archivo de log E2E."""
    LOG.append(msg)


def _norm(value: object) -> str | float | int:
    """Normaliza valor para comparación: None→vacío, float redondeado."""
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 2)
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    return str(value).strip()


def _norm_compare(value: object) -> str:
    """Normaliza valor a string comparable para diff de workbooks."""
    normalized = _norm(value)
    if isinstance(normalized, int):
        return str(normalized)
    if isinstance(normalized, float):
        return str(normalized)
    return str(normalized)


def compare_workbooks(generated: Path, reference: Path) -> dict:
    """Compara filas y encabezados entre Excel generado y referencia."""
    from openpyxl import load_workbook

    from legalizacion_tc.template_layout import detect_template_layout

    def read_data_rows(path: Path) -> dict:
        """Lee filas de datos, encabezado y totales de un workbook de legalización."""
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        layout = detect_template_layout(ws)
        rows: list[dict] = []
        for row_idx in range(19, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=7).value
            if isinstance(label, str) and label.strip().lower() == "valor total":
                break
            vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 13)]
            if any(v is not None and str(v).strip() != "" for v in vals):
                total_col = layout.col_total or layout.col_cops
                rows.append(
                    {
                        "row": row_idx,
                        "numero_factura": vals[0],
                        "nit": vals[1],
                        "razon": vals[2],
                        "detalle": vals[3],
                        "articulo": vals[4],
                        "centro": vals[5],
                        "moneda": vals[6],
                        "usd": ws.cell(row=row_idx, column=layout.col_usd).value,
                        "cops": ws.cell(row=row_idx, column=layout.col_cops).value,
                        "total": ws.cell(row=row_idx, column=total_col).value,
                    }
                )
        header = {
            "B7": ws["B7"].value,
            "B9": ws["B9"].value,
            "B11": str(ws["B11"].value or "").strip(),
            "B13": ws["B13"].value,
        }
        footer: dict = {}
        for row_idx in range(19, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=7).value
            if isinstance(label, str) and label.strip().lower() == "valor total":
                footer["totals_row"] = row_idx
                footer["sum_usd"] = ws.cell(row=row_idx, column=layout.col_usd).value
                footer["sum_cops"] = ws.cell(row=row_idx, column=layout.col_cops).value
                total_col = layout.col_total or layout.col_cops
                footer["sum_total"] = ws.cell(row=row_idx, column=total_col).value
        wb.close()
        return {"header": header, "rows": rows, "footer": footer}

    gen = read_data_rows(generated)
    ref = read_data_rows(reference)
    diffs: list[str] = []
    if gen["header"]["B11"] != ref["header"]["B11"]:
        diffs.append(f"B11 solicitante: gen={gen['header']['B11']} ref={ref['header']['B11']}")
    if len(gen["rows"]) != len(ref["rows"]):
        diffs.append(
            f"Cantidad filas: gen={len(gen['rows'])} ref={len(ref['rows'])} (extracto incompleto)"
        )
    for i, (g, r) in enumerate(zip(gen["rows"], ref["rows"], strict=False)):
        for key in ("numero_factura", "nit", "razon", "articulo", "centro", "moneda", "usd", "cops", "total"):
            gv, rv = _norm_compare(g.get(key)), _norm_compare(r.get(key))
            if gv != rv:
                diffs.append(f"Fila {i+1} {key}: gen={gv} ref={rv}")
    return {"generated": gen, "reference": ref, "diffs": diffs, "match": len(diffs) == 0}


def main() -> int:
    """Punto de entrada CLI del script."""
    parser = argparse.ArgumentParser(description="E2E legalización TC (carpeta local)")
    add_folder_argument(parser)
    parser.add_argument(
        "--require-invoices",
        action="store_true",
        help="Exigir JSON de facturas en .cache/invoices/ (por defecto se omite la validación).",
    )
    args = parser.parse_args()

    result: dict = {"ok": False, "log": LOG, "folder": None}
    try:
        folder = resolve_e2e_folder(args.folder)
        result["folder"] = str(folder)
        log(f"Carpeta: {folder.name}")
        log(f"Archivos: {[f.name for f in sorted(folder.iterdir())]}")

        from legalizacion_tc.run_pipeline import run_pipeline

        outcomes = run_pipeline(
            local_folder=folder,
            skip_invoice_check=not args.require_invoices,
            dry_run=False,
        )
        if len(outcomes) != 1:
            raise RuntimeError(
                f"Se esperaba una sola tarjeta en {folder.name}; detectadas: {len(outcomes)}"
            )
        outcome = outcomes[0]
        if outcome.error is not None:
            raise outcome.error
        pipeline_result = outcome.result
        if pipeline_result is None:
            raise RuntimeError("Pipeline sin resultado")
        log(f"Output: {pipeline_result.output_path}")
        ok_matches = [m for m in pipeline_result.matches if m.status in {"OK", "GMF"}]
        log(f"Matches OK: {len(ok_matches)}/{len(pipeline_result.matches)}")

        out_path = Path(pipeline_result.output_path)
        from legalizacion_tc.reference_loader import find_manual_reference_for_compare

        reference = find_manual_reference_for_compare(folder, pipeline_result.extract.card)
        comparison = None
        if reference and out_path.exists():
            comparison = compare_workbooks(out_path, reference)
            log(f"Comparación: match={comparison['match']}, diffs={comparison['diffs']}")
        else:
            log(f"Referencia no encontrada o output ausente: ref={reference}, out={out_path.exists()}")

        result.update(
            {
                "ok": True,
                "output_path": str(out_path),
                "reference_path": str(reference) if reference else None,
                "card": pipeline_result.extract.card,
                "matches": [
                    {
                        "desc": m.transaction.description,
                        "amount": m.transaction.amount_cop,
                        "status": m.status,
                        "invoice": m.invoice.source_filename if m.invoice else None,
                    }
                    for m in pipeline_result.matches
                ],
                "comparison": comparison,
            }
        )
    except Exception:
        log(traceback.format_exc())
        result["error"] = traceback.format_exc()
    result["log"] = LOG
    RESULT.write_text(json.dumps(result, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    if not result.get("ok"):
        return 1
    comparison = result.get("comparison") or {}
    return 0 if comparison.get("match") else 1


if __name__ == "__main__":
    raise SystemExit(main())
