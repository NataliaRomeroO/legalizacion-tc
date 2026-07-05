"""Generación y merge del Excel ``Formato de Legalización TC`` desde plantilla.

- ``build_legalization_workbook``: nuevo archivo desde ``plantilla_base.xlsx``.
- ``merge_legalization_workbook``: append sobre Formato base existente (dedup previo).
- Filas desde fila 19; resaltado amarillo (``FFFF00``) si ``needs_review``.
- Footer: SUM, total extractos, checkpoint = extractos − total (debe ≈ 0).
- ``relabel_existing_rows``: etiquetas ``Legalizado en {mes} corte N``.
"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings
from .legalization_batch import (
    current_batch_label,
    execution_period_month,
    relabel_existing_rows,
)
from .models import CardMetadata, ExtractData, LegalizationRow
from .template_layout import TemplateLayout, detect_template_layout, total_sum_column

DATA_START_ROW = 19
HEADER_ROW = 18
FOOTER_LABEL = "Valor total"
_REVIEW_FILL = PatternFill("solid", fgColor="FFFF00")


def _set_cell_value(ws: Worksheet, row_idx: int, column: int, value) -> None:
    """Asigna valor vía ``.value`` para reemplazar fórmulas obsoletas de la plantilla."""
    ws.cell(row=row_idx, column=column).value = value


def create_minimal_template(path: Path) -> None:
    """Genera plantilla mínima local cuando no hay ``PLANTILLA_DRIVE_FILE_ID`` configurado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato"
    ws["A5"] = "Información general de la solicitud"
    ws["A7"], ws["B7"] = "Fecha de legalización:", None
    ws["A9"], ws["B9"] = "Solicitud para:", None
    ws["A11"], ws["B11"] = "Solicitante:", None
    ws["A13"], ws["B13"] = "Concepto de legalización:", None
    ws["A17"] = "Información detallada de la solicitud"
    headers = [
        "Nº Factura",
        "Nit Proveedor",
        "Razon social",
        "Detalle del Gasto",
        "Articulo Contable",
        "Centro de costo",
        "Moneda",
        "Valor base de la compra en USD",
        "Valor base de la compra en CLP",
        "Valor base de la compra en SOL: ",
        "Valor base de la compra en COPS",
        "Iva de la compa",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    footer_row = DATA_START_ROW + 1
    ws.cell(row=footer_row, column=7, value=FOOTER_LABEL)
    ws.cell(row=footer_row + 3, column=7, value="Extractos o movimientos")
    ws.cell(row=footer_row + 4, column=7, value="Checkpoint")
    wb.save(path)
    wb.close()


def _find_footer_start(ws: Worksheet) -> int:
    """Primera fila con etiqueta 'Valor total' (columna G) o max_row+1."""
    for row_idx in range(DATA_START_ROW, ws.max_row + 2):
        label = ws.cell(row=row_idx, column=7).value
        if isinstance(label, str) and label.strip().lower() == FOOTER_LABEL.lower():
            return row_idx
    return ws.max_row + 1


def _clear_data_row_values(ws: Worksheet, row_idx: int, max_col: int = 12) -> None:
    """Limpia celdas de una fila de datos sin alterar estilos."""
    for col in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col).value = None


def _ensure_data_capacity(ws: Worksheet, needed_rows: int) -> int:
    """Inserta o limpia filas entre datos y footer para ajustar capacidad; retorna footer_start."""
    footer_start = _find_footer_start(ws)
    available = footer_start - DATA_START_ROW
    if needed_rows > available:
        ws.insert_rows(footer_start, needed_rows - available)
    elif needed_rows < available:
        for row_idx in range(DATA_START_ROW + needed_rows, footer_start):
            _clear_data_row_values(ws, row_idx)
    return _find_footer_start(ws)


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int, max_col: int = 12) -> None:
    """Copia fuente, borde, relleno y formato numérico de una fila plantilla a otra."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _max_style_col(layout: TemplateLayout) -> int:
    """Última columna a estilizar/resaltar según columnas opcionales del layout."""
    return layout.col_legalizado or layout.col_soporte or layout.col_total or layout.col_iva


def _write_row(
    ws: Worksheet,
    row_idx: int,
    row: LegalizationRow,
    layout: TemplateLayout,
    batch_label: str | None = None,
) -> None:
    """Escribe una fila de legalización respetando columnas detectadas del layout."""
    ws.cell(row=row_idx, column=1, value=row.numero_factura or None)
    ws.cell(row=row_idx, column=2, value=row.nit_proveedor or None)
    ws.cell(row=row_idx, column=3, value=row.razon_social or None)
    ws.cell(row=row_idx, column=4, value=row.detalle_gasto or None)
    ws.cell(row=row_idx, column=5, value=row.articulo_contable or None)
    ws.cell(row=row_idx, column=6, value=row.centro_costo or None)
    ws.cell(row=row_idx, column=7, value=row.moneda or None)
    _set_cell_value(ws, row_idx, layout.col_usd, row.valor_base_usd)
    _set_cell_value(ws, row_idx, layout.col_clp, row.valor_base_clp)
    if layout.col_sol is not None:
        _set_cell_value(ws, row_idx, layout.col_sol, row.valor_base_sol)
    _set_cell_value(ws, row_idx, layout.col_cops, row.valor_base_cops)
    iva_cell = ws.cell(row=row_idx, column=layout.col_iva)
    iva_cell.value = row.iva if row.iva else None
    if (
        layout.col_total
        and layout.col_total != layout.col_cops
        and layout.col_total != layout.col_iva
    ):
        _set_cell_value(ws, row_idx, layout.col_total, row.valor_total_compra_cop)


    if layout.col_soporte:
        ws.cell(
            row=row_idx,
            column=layout.col_soporte,
            value=row.documento_soporte or None,
        )
    if layout.col_legalizado and batch_label:
        ws.cell(row=row_idx, column=layout.col_legalizado, value=batch_label)


def _apply_review_highlight(
    ws: Worksheet, row_idx: int, layout: TemplateLayout, row: LegalizationRow
) -> None:
    """Resalta fila en amarillo (``FFFF00``) cuando ``needs_review`` es True."""
    if not row.needs_review:
        return
    max_col = _max_style_col(layout)
    for col in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col).fill = _REVIEW_FILL


def _column_letter(col: int) -> str:
    """Convierte índice de columna 1-based a letra Excel (A, B, …, AA)."""
    letters = ""
    while col:
        col, remainder = divmod(col - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _is_extractos_checkpoint_label(value) -> bool:
    """True si el valor es etiqueta de pie 'Extractos o movimientos' o 'Checkpoint'."""
    if not isinstance(value, str):
        return False
    lower = value.strip().lower()
    return lower == "extractos o movimientos" or lower == "checkpoint"


def _footer_label_col(layout: TemplateLayout) -> int:
    """Columna de etiquetas del bloque extractos/checkpoint según layout."""
    return layout.extractos_label_col or layout.checkpoint_label_col or 9


def _footer_value_col(layout: TemplateLayout) -> int:
    """Columna de valores del bloque extractos/checkpoint según layout."""
    return layout.extractos_value_col or layout.checkpoint_value_col or total_sum_column(
        layout
    )


def _clear_obsolete_footer_blocks(
    ws: Worksheet,
    footer_start: int,
    active_extractos_row: int,
    active_checkpoint_row: int,
    label_col: int,
    value_col: int,
) -> None:
    """Elimina bloques duplicados de extractos/checkpoint tras merges sucesivos."""
    active = {active_extractos_row, active_checkpoint_row}
    value_cols = {value_col, 10, 12}

    for row_idx in range(footer_start + 1, footer_start + 21):
        if row_idx in active:
            continue

        stale_label = False
        for col in (7, 9, label_col):
            if _is_extractos_checkpoint_label(ws.cell(row=row_idx, column=col).value):
                ws.cell(row=row_idx, column=col).value = None
                stale_label = True

        is_gap = row_idx in (footer_start + 1, footer_start + 2)
        if is_gap or stale_label:
            ws.cell(row=row_idx, column=label_col).value = None
            for col in value_cols:
                ws.cell(row=row_idx, column=col).value = None


def _update_footer_formulas(
    ws: Worksheet,
    last_data_row: int,
    total_extract_cop: float,
    layout: TemplateLayout,
) -> None:
    """Actualiza SUM por columna, total extractos y checkpoint (extractos − total ≈ 0)."""
    footer_start = _find_footer_start(ws)
    totals_row = footer_start
    ws.cell(row=totals_row, column=7, value=FOOTER_LABEL)

    for col in layout.sum_columns:
        col_letter = _column_letter(col)
        ws.cell(
            row=totals_row,
            column=col,
            value=f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{last_data_row})",
        )

    label_col = _footer_label_col(layout)
    extractos_value_col = _footer_value_col(layout)
    checkpoint_value_col = layout.checkpoint_value_col or extractos_value_col

    extractos_row = footer_start + 3
    checkpoint_row = footer_start + 4

    _clear_obsolete_footer_blocks(
        ws,
        footer_start,
        extractos_row,
        checkpoint_row,
        label_col,
        extractos_value_col,
    )

    ws.cell(row=extractos_row, column=label_col, value="Extractos o movimientos")
    ws.cell(row=extractos_row, column=extractos_value_col, value=total_extract_cop)

    total_col = total_sum_column(layout)
    total_letter = _column_letter(total_col)
    extractos_letter = _column_letter(extractos_value_col)
    ws.cell(row=checkpoint_row, column=label_col, value="Checkpoint")
    ws.cell(
        row=checkpoint_row,
        column=checkpoint_value_col,
        value=f"={extractos_letter}{extractos_row}-{total_letter}{totals_row}",
    )


def _sanitize_layout_columns(layout: TemplateLayout) -> TemplateLayout:
    """Evita col_total == col_iva (plantillas antiguas); anula col_total en ese caso."""
    if layout.col_total == layout.col_iva:
        layout.col_total = None
    return layout


def output_filename(extract: ExtractData) -> str:
    """Nombre estándar del archivo de salida: tarjeta, mes y año del extracto."""
    return (
        f"Formato de Legalización TC {extract.card} - "
        f"{extract.period_month} - {extract.period_year}.xlsx"
    )


def _count_existing_data_rows(ws: Worksheet) -> int:
    """Cuenta filas con detalle (columna D) entre DATA_START_ROW y el footer."""
    footer_start = _find_footer_start(ws)
    count = 0
    for row_idx in range(DATA_START_ROW, footer_start):
        if ws.cell(row=row_idx, column=4).value not in (None, ""):
            count += 1
    return count


def _read_existing_batch_labels(
    ws: Worksheet, layout: TemplateLayout, existing_rows: int
) -> list[str | None]:
    """Lee etiquetas 'Legalizado en …' de filas existentes; None si columna ausente."""
    if not layout.col_legalizado:
        return [None] * existing_rows
    labels: list[str | None] = []
    for offset in range(existing_rows):
        row_idx = DATA_START_ROW + offset
        value = ws.cell(row=row_idx, column=layout.col_legalizado).value
        labels.append(str(value).strip() if value not in (None, "") else None)
    return labels


def _apply_header_metadata(
    ws: Worksheet,
    settings: Settings,
    extract: ExtractData,
    card_meta: CardMetadata,
) -> None:
    """Rellena encabezado: fecha, solicitante, concepto y centro de costo."""
    ws["B7"] = date.today()
    ws["B9"] = settings.solicitud_para
    ws["B11"] = card_meta.solicitante
    ws["B13"] = (
        f"LEGALIZACIÓN TC {extract.card} - {extract.period_month} {extract.period_year}"
    )


def merge_legalization_workbook(
    settings: Settings,
    existing_path: Path,
    output_path: Path,
    extract: ExtractData,
    card_meta: CardMetadata,
    new_rows: list[LegalizationRow],
) -> Path:
    """Append sobre Formato base existente: relabel filas previas, nuevas filas y footer.

    Copia ``existing_path`` a ``output_path`` si difieren. Inserta filas antes del footer
    cuando hace falta. Si ``new_rows`` está vacío, solo actualiza metadatos y fórmulas.
    """
    if existing_path.resolve() != output_path.resolve():
        shutil.copy2(existing_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    _apply_header_metadata(ws, settings, extract, card_meta)

    layout = _sanitize_layout_columns(detect_template_layout(ws))
    execution_month = execution_period_month(timezone=settings.timezone)
    existing_count = _count_existing_data_rows(ws)
    existing_labels = _read_existing_batch_labels(ws, layout, existing_count)
    updated_labels = relabel_existing_rows(existing_labels, execution_month)
    for offset, label in enumerate(updated_labels):
        if layout.col_legalizado and label:
            ws.cell(row=DATA_START_ROW + offset, column=layout.col_legalizado, value=label)

    footer_start = _find_footer_start(ws)
    if new_rows:
        last_data_row = DATA_START_ROW + existing_count - 1 if existing_count else DATA_START_ROW - 1
        first_append_row = last_data_row + 1
        available_empty = max(0, footer_start - first_append_row)
        if len(new_rows) > available_empty:
            ws.insert_rows(footer_start, len(new_rows) - available_empty)
            footer_start = _find_footer_start(ws)

        style_source = DATA_START_ROW
        batch_label = current_batch_label(execution_month)
        for offset, row in enumerate(new_rows):
            target_row = first_append_row + offset
            _copy_row_style(ws, style_source, target_row, max_col=_max_style_col(layout))
            _write_row(ws, target_row, row, layout, batch_label=batch_label)
            _apply_review_highlight(ws, target_row, layout, row)

    total_rows = existing_count + len(new_rows)
    last_data_row = DATA_START_ROW + total_rows - 1 if total_rows else DATA_START_ROW - 1
    if last_data_row >= DATA_START_ROW:
        _update_footer_formulas(ws, last_data_row, extract.total_cop, layout)

    wb.save(output_path)
    wb.close()
    return output_path


def build_legalization_workbook(
    settings: Settings,
    template_path: Path,
    output_path: Path,
    extract: ExtractData,
    card_meta: CardMetadata,
    rows: list[LegalizationRow],
    *,
    existing_workbook_path: Path | None = None,
) -> Path:
    """Escribe Excel de legalización; delega en merge si ``existing_workbook_path`` existe.

    Creación desde plantilla: copia segura si origen y destino son el mismo path.
    """
    if existing_workbook_path and existing_workbook_path.exists():
        return merge_legalization_workbook(
            settings,
            existing_workbook_path,
            output_path,
            extract,
            card_meta,
            rows,
        )

    source = template_path
    if source.resolve() == output_path.resolve():
        tmp = output_path.with_suffix(".tmp.xlsx")
        shutil.copy2(source, tmp)
        source = tmp
    if output_path.exists():
        output_path.unlink()
    shutil.copy2(source, output_path)
    if source != template_path:
        source.unlink(missing_ok=True)
    wb = load_workbook(output_path)

    ws = wb.active
    _apply_header_metadata(ws, settings, extract, card_meta)

    _ensure_data_capacity(ws, len(rows))
    layout = _sanitize_layout_columns(detect_template_layout(ws))
    execution_month = execution_period_month(timezone=settings.timezone)
    style_source = DATA_START_ROW
    batch_label = current_batch_label(execution_month)
    for offset, row in enumerate(rows):
        target_row = DATA_START_ROW + offset
        _copy_row_style(ws, style_source, target_row, max_col=_max_style_col(layout))
        _write_row(ws, target_row, row, layout, batch_label=batch_label)
        _apply_review_highlight(ws, target_row, layout, row)

    last_data_row = DATA_START_ROW + len(rows) - 1 if rows else DATA_START_ROW - 1
    if last_data_row >= DATA_START_ROW:
        _update_footer_formulas(ws, last_data_row, extract.total_cop, layout)

    wb.save(output_path)
    wb.close()
    return output_path
