"""Write-back de resultados de conciliación al preliminar Excel.

Añade o rellena columnas ``Validación`` (OK/REVISAR/NO) y ``Observaciones`` por
``row_index`` del movimiento. Solo aplica cuando el origen de movimientos es Excel,
no PDF.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import MatchResult
from .validation_notes import validation_flag

VALIDATION_COLUMN = "Validación"
OBSERVATIONS_COLUMN = "Observaciones"


def _find_or_create_column(ws, header: str, after_col: int | None = None) -> int:
    """Busca columna por encabezado en fila 1; la crea al final o tras ``after_col``."""
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and str(cell_value).strip() == header:
            return col_idx
    if after_col is not None:
        insert_at = after_col + 1
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value=header)
        return insert_at
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=header)
    return new_col


def apply_extract_review_columns(extract_path: Path, matches: list[MatchResult]) -> Path:
    """Escribe columnas Validación y Observaciones en el preliminar Excel por ``row_index``."""
    wb = load_workbook(extract_path)
    ws = wb.active

    validation_col = _find_or_create_column(ws, VALIDATION_COLUMN)
    observations_col = _find_or_create_column(ws, OBSERVATIONS_COLUMN, after_col=validation_col)

    col_letter = get_column_letter(observations_col)
    ws.column_dimensions[col_letter].width = 60

    match_by_row = {m.transaction.row_index: m for m in matches}
    for row_idx in range(2, ws.max_row + 1):
        match = match_by_row.get(row_idx)
        if not match:
            continue
        ws.cell(row=row_idx, column=validation_col, value=validation_flag(match))
        ws.cell(row=row_idx, column=observations_col, value=match.observacion or "")

    wb.save(extract_path)
    wb.close()
    return extract_path


def apply_validation_column(extract_path: Path, matches: list[MatchResult]) -> Path:
    """Alias retrocompatible."""
    return apply_extract_review_columns(extract_path, matches)
