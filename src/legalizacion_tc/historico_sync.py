"""Utilidad offline: escanea Formatos locales y construye histórico de proveedores.

``scan_local_legalizations`` hace rglob de ``Formato de Legalización*.xlsx``;
última fila por NIT gana. Asume layout clásico de 12 columnas (sin detección de plantilla).
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import LegalizationRow, ProviderHistory
from .nit_utils import index_historico

LEGALIZATION_PATTERN = re.compile(r"^Formato de Legalizaci", re.IGNORECASE)
DATA_START_ROW = 19
HEADERS = [
    "Nº Factura",
    "Nit Proveedor",
    "Razon social",
    "Detalle del Gasto",
    "Articulo Contable",
    "Centro de costo",
    "Moneda",
    "Valor base de la compra en USD",
    "Valor base de la compra en COPS",
    "Iva de la compa",
    "Valor total compra",
    "Documento Soporte",
]


def _parse_rows_from_workbook(path: Path) -> list[LegalizationRow]:
    """Extrae filas de datos de un Formato local con layout clásico de 12 columnas."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[LegalizationRow] = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        factura = ws.cell(row=row_idx, column=1).value
        if not factura:
            continue
        label_g = ws.cell(row=row_idx, column=7).value
        if isinstance(label_g, str) and label_g.strip().lower() == "valor total":
            break
        rows.append(
            LegalizationRow(
                numero_factura=str(factura),
                nit_proveedor=str(ws.cell(row=row_idx, column=2).value or ""),
                razon_social=str(ws.cell(row=row_idx, column=3).value or ""),
                detalle_gasto=str(ws.cell(row=row_idx, column=4).value or ""),
                articulo_contable=str(ws.cell(row=row_idx, column=5).value or ""),
                centro_costo=str(ws.cell(row=row_idx, column=6).value or ""),
                moneda=str(ws.cell(row=row_idx, column=7).value or ""),
                valor_base_usd=_to_float(ws.cell(row=row_idx, column=8).value),
                valor_base_cops=_to_float(ws.cell(row=row_idx, column=9).value) or 0.0,
                iva=_to_float(ws.cell(row=row_idx, column=10).value) or 0.0,
                valor_total_compra_cop=_to_float(ws.cell(row=row_idx, column=11).value) or 0.0,
                documento_soporte=str(ws.cell(row=row_idx, column=12).value or "NO"),
            )
        )
    wb.close()
    return rows


def _to_float(value: object) -> float | None:
    """Convierte celda Excel a float; devuelve ``None`` si no es numérico."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def scan_local_legalizations(folder: Path) -> dict[str, ProviderHistory]:
    """Escanea Formatos locales y construye histórico; la última fila por NIT prevalece."""
    historico: dict[str, ProviderHistory] = {}
    for path in folder.rglob("*.xlsx"):
        if not LEGALIZATION_PATTERN.search(path.name):
            continue
        for row in _parse_rows_from_workbook(path):
            if not row.nit_proveedor:
                continue
            historico[row.nit_proveedor] = ProviderHistory(
                nit=row.nit_proveedor,
                razon_social=row.razon_social,
                detalle_gasto=row.detalle_gasto,
                articulo_contable=row.articulo_contable,
                fecha_ultima="",
                archivo_origen=path.name,
            )
    return index_historico(historico)
