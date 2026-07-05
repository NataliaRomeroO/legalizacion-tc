"""Deduplicación de movimientos ya legalizados en Formatos Excel previos (re-ejecución segura).

Estrategias en ``is_match_already_legalized``:
- Huella (fecha + monto + concepto parseado del detalle).
- Clave fecha+monto (cargos consolidados mismo día).
- Par (numero_factura, nit); si NIT vacío, solo número de factura.
- GMF: por monto individual o suma de GMF previos (PDF puede consolidar).

``filter_matches_for_append`` retorna matches nuevos + contador de omitidos.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .date_normalizer import parse_detalle_date_suffix
from .excel_report_builder import DATA_START_ROW, FOOTER_LABEL
from .metadata_mapper import _format_numero_factura
from .models import InvoiceData, MatchResult, Transaction
from .nit_utils import normalize_nit_key
from .template_layout import TemplateLayout, detect_template_layout, total_sum_column

DATE_SUFFIX_RE = re.compile(
    r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})\s*$"
)
_SPANISH_SUFFIX_STRIP_RE = re.compile(
    r"(\d{1,2})\s+DE\s+\w+\s*$",
    re.IGNORECASE,
)
_PERIOD_YEAR_RE = re.compile(r"(20\d{2})")


def normalize_desc(desc: str) -> str:
    """Normaliza descripción para huellas: trim, mayúsculas y espacios colapsados."""
    return re.sub(r"\s+", " ", (desc or "").strip().upper())


_EXCEL_FLOAT_STR_RE = re.compile(r"^\d+\.0+$")


def _excel_scalar_text(value: object) -> str | None:
    """Convierte valor de celda Excel a texto sin artefactos float (p. ej. 53156.0 → '53156')."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text or None


def normalize_invoice_key(value: object) -> str:
    """Clave canónica de número de factura desde Excel o JSON (enteros, floats .0, vacíos)."""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value).upper()
    if isinstance(value, (int, float)):
        return str(int(value))
    text = str(value).strip().upper()
    # Excel may serialize integers as "233888670.0"
    if _EXCEL_FLOAT_STR_RE.fullmatch(text):
        return text.split(".", 1)[0]
    return text


def transaction_fingerprint(tx: Transaction) -> tuple[str, float, str]:
    """Huella (fecha ISO, monto COP redondeado, descripción normalizada) de un movimiento."""
    return (
        tx.tx_date.isoformat(),
        round(tx.amount_cop, 2),
        normalize_desc(tx.description),
    )


def _parse_detalle_parts(
    detalle: str, default_year: int | None = None
) -> tuple[str, date | None]:
    """Extrae concepto y fecha del detalle Excel (prefijo TC, sufijos dd/mm o 'DE MES')."""
    text = str(detalle or "").strip()
    tx_date = parse_detalle_date_suffix(text, default_year=default_year)
    body = text
    if tx_date is not None:
        body = _SPANISH_SUFFIX_STRIP_RE.sub("", body).strip()
        body = DATE_SUFFIX_RE.sub("", body).strip()
    else:
        body = DATE_SUFFIX_RE.sub("", body).strip()
    match = re.match(r"^TC\s+\d+\s+(.+)$", body, re.IGNORECASE)
    desc = match.group(1).strip() if match else body
    return normalize_desc(desc), tx_date


def _cell_amount(value) -> float | None:
    """Parsea monto numérico de celda (int, float o texto con $ y comas)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def _infer_period_year_from_workbook(ws, path: Path) -> int:
    """Infiere año del periodo desde encabezado del libro o nombre de archivo; fallback año actual."""
    for row_idx in range(1, 20):
        for col_idx in (1, 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            match = _PERIOD_YEAR_RE.search(str(value))
            if match:
                return int(match.group(1))
    match = _PERIOD_YEAR_RE.search(path.name)
    if match:
        return int(match.group(1))
    return date.today().year


def invoice_nit_key(
    numero: str | None,
    nit: str | None,
    invoice: InvoiceData | None = None,
) -> tuple[str, str] | None:
    """Par (número factura, NIT normalizado) o None si falta número de factura."""
    formatted = _format_numero_factura(numero, invoice)
    num_key = normalize_invoice_key(formatted)
    nit_key = normalize_nit_key(nit)
    if not num_key:
        return None
    return (num_key, nit_key)


def _invoices_from_match(match: MatchResult) -> list[InvoiceData]:
    """Lista facturas asociadas al match: principal, secundaria y componentes."""
    invoices: list[InvoiceData] = []
    if match.invoice is not None:
        invoices.append(match.invoice)
    if match.secondary_invoice is not None:
        invoices.append(match.secondary_invoice)
    if match.component_invoices:
        invoices.extend(match.component_invoices)
    return invoices


@dataclass
class LegalizedState:
    """Estado acumulado de movimientos ya presentes en Formatos Excel previos."""

    fingerprints: set[tuple[str, float, str]] = field(default_factory=set)
    date_amount_keys: set[tuple[str, float]] = field(default_factory=set)
    invoice_nit_keys: set[tuple[str, str]] = field(default_factory=set)
    gmf_amount_keys: set[float] = field(default_factory=set)


def _find_footer_start(ws) -> int:
    """Primera fila del pie ('Valor total') o max_row+1 si no hay footer."""
    for row_idx in range(DATA_START_ROW, ws.max_row + 2):
        label = ws.cell(row=row_idx, column=7).value
        if isinstance(label, str) and label.strip().lower() == FOOTER_LABEL.lower():
            return row_idx
    return ws.max_row + 1


def _row_cop_amount(ws, row_idx: int, layout: TemplateLayout) -> float | None:
    """Monto COP de fila: columna total del layout o fallback a col_cops."""
    total_col = layout.col_total or total_sum_column(layout)
    amount = _cell_amount(ws.cell(row=row_idx, column=total_col).value)
    if amount is not None:
        return amount
    return _cell_amount(ws.cell(row=row_idx, column=layout.col_cops).value)


def merge_legalized_states(*states: LegalizedState) -> LegalizedState:
    """Une varios estados (varios Formatos) en uno solo para dedup global."""
    merged = LegalizedState()
    for state in states:
        merged.fingerprints.update(state.fingerprints)
        merged.date_amount_keys.update(state.date_amount_keys)
        merged.invoice_nit_keys.update(state.invoice_nit_keys)
        merged.gmf_amount_keys.update(state.gmf_amount_keys)
    return merged


def legalized_state_from_workbook(
    path: Path,
    *,
    default_year: int | None = None,
) -> LegalizedState:
    """Construye estado desde un Formato Excel; vacío si el archivo no existe.

    Registra huellas, claves fecha+monto (incl. suma mismo día), par factura+NIT
    y montos GMF sin fecha. Si hay varios cargos el mismo día, agrega clave con total.
    """
    state = LegalizedState()
    if not path.exists():
        return state

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    layout = detect_template_layout(ws)
    footer_start = _find_footer_start(ws)
    period_year = default_year or _infer_period_year_from_workbook(ws, path)

    rows_by_date: dict[str, list[tuple[str, float]]] = {}

    for row_idx in range(DATA_START_ROW, footer_start):
        detalle = ws.cell(row=row_idx, column=4).value
        if detalle is None or str(detalle).strip() == "":
            continue
        detalle_text = str(detalle)

        numero = _excel_scalar_text(ws.cell(row=row_idx, column=1).value)
        nit = _excel_scalar_text(ws.cell(row=row_idx, column=2).value)
        inv_key = invoice_nit_key(numero, nit)
        if inv_key is not None:
            state.invoice_nit_keys.add(inv_key)

        desc, tx_date = _parse_detalle_parts(detalle_text, default_year=period_year)
        amount = _row_cop_amount(ws, row_idx, layout)
        if amount is None:
            continue
        if tx_date is None and "GMF" in detalle_text.upper():
            state.gmf_amount_keys.add(round(amount, 2))
            continue
        if tx_date is None:
            continue

        iso = tx_date.isoformat()
        rounded = round(amount, 2)
        state.fingerprints.add((iso, rounded, desc))
        state.date_amount_keys.add((iso, rounded))
        rows_by_date.setdefault(iso, []).append((desc, rounded))

    for iso, entries in rows_by_date.items():
        if len(entries) < 2:
            continue
        total = round(sum(amount for _, amount in entries), 2)
        state.date_amount_keys.add((iso, total))

    wb.close()
    return state


def legalized_state_from_paths(
    paths: list[Path],
    *,
    default_year: int | None = None,
) -> LegalizedState:
    """Agrega estados de varios archivos Excel mediante ``merge_legalized_states``."""
    states = [
        legalized_state_from_workbook(path, default_year=default_year)
        for path in paths
    ]
    return merge_legalized_states(*states)


def is_transaction_fingerprint_legalized(
    tx: Transaction, state: LegalizedState
) -> bool:
    """True si la huella o la clave (fecha, monto) ya está en el estado."""
    fp = transaction_fingerprint(tx)
    if fp in state.fingerprints:
        return True
    key = (tx.tx_date.isoformat(), round(tx.amount_cop, 2))
    return key in state.date_amount_keys


def is_transaction_legalized(tx: Transaction, state: LegalizedState) -> bool:
    """Alias de ``is_transaction_fingerprint_legalized`` para movimientos sueltos."""
    return is_transaction_fingerprint_legalized(tx, state)


def _invoice_key_already_legalized(
    key: tuple[str, str], state: LegalizedState
) -> bool:
    """True si el par factura+NIT coincide; con NIT vacío en cualquier lado, solo número."""
    if key in state.invoice_nit_keys:
        return True
    num_key, nit_key = key
    # Manual Formatos may have NIT filled from historico while the invoice JSON
    # has null NIT (common for foreign vendors). Match on invoice number alone
    # when either side is missing NIT.
    for prior_num, prior_nit in state.invoice_nit_keys:
        if prior_num != num_key:
            continue
        if not nit_key or not prior_nit:
            return True
    return False


def is_match_already_legalized(match: MatchResult, state: LegalizedState) -> bool:
    """Determina si un match reconciliado ya fue legalizado en ejecuciones previas.

    GMF: huella, monto individual en ``gmf_amount_keys``, o suma de GMF previos
    (PDF puede consolidar varias líneas GMF en un solo movimiento).
    OK con factura: par factura+NIT (NIT flexible) y luego huella del movimiento.
    Otros estados: solo huella del movimiento.
    """
    if match.status == "GMF":
        if is_transaction_fingerprint_legalized(match.transaction, state):
            return True
        amount = round(match.transaction.amount_cop, 2)
        if amount in state.gmf_amount_keys:
            return True
        # PDF may consolidate several GMF lines into one movement.
        prior_gmf_total = round(sum(state.gmf_amount_keys), 2)
        return prior_gmf_total > 0 and amount == prior_gmf_total

    if match.status == "OK" and match.invoice is not None:
        for invoice in _invoices_from_match(match):
            key = invoice_nit_key(
                invoice.numero_factura,
                invoice.nit_proveedor,
                invoice,
            )
            if key is not None and _invoice_key_already_legalized(key, state):
                return True
        return is_transaction_fingerprint_legalized(match.transaction, state)

    return is_transaction_fingerprint_legalized(match.transaction, state)


def filter_matches_for_append(
    matches: list[MatchResult], state: LegalizedState
) -> tuple[list[MatchResult], int]:
    """Filtra matches para append: retorna (nuevos, omitidos) usando ``LegalizedState``.

    Los omitidos son re-ejecuciones seguras sobre el mismo periodo (sin duplicar filas).
    """
    kept: list[MatchResult] = []
    skipped = 0
    for match in matches:
        if is_match_already_legalized(match, state):
            skipped += 1
            continue
        kept.append(match)
    return kept, skipped
