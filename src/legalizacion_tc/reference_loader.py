"""Histórico de proveedores desde Excel de referencia manual (modo local/demo).

``find_reference_legalization`` prefiere legacy manual ``MAYO-2026`` sobre generado.
``load_historico_from_reference`` lee NIT, razón, detalle y artículo de filas de datos.
Solo usado cuando ``run_pipeline`` recibe ``--local-folder``.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import ProviderHistory
from .nit_utils import index_historico

DATA_START_ROW = 19
FOOTER_LABEL = "valor total"
GENERATED_NAME_PATTERN = re.compile(r"MAYO\s+-\s+\d{4}", re.IGNORECASE)
MANUAL_NAME_PATTERN = re.compile(r"MAYO-\s*\d{4}", re.IGNORECASE)


def load_historico_from_reference(path: Path) -> dict[str, ProviderHistory]:
    """Lee histórico de proveedores desde un Excel de referencia local (modo demo)."""
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    mapping: dict[str, ProviderHistory] = {}
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=7).value
        if isinstance(label, str) and label.strip().lower() == FOOTER_LABEL:
            break
        nit = ws.cell(row=row_idx, column=2).value
        if nit is None or str(nit).strip() == "":
            continue
        nit_key = str(nit).strip()
        mapping[nit_key] = ProviderHistory(
            nit=nit_key,
            razon_social=str(ws.cell(row=row_idx, column=3).value or "").strip(),
            detalle_gasto=str(ws.cell(row=row_idx, column=4).value or "").strip(),
            articulo_contable=str(ws.cell(row=row_idx, column=5).value or "").strip(),
            archivo_origen=path.name,
        )
    wb.close()
    return index_historico(mapping)


def find_reference_legalization(folder: Path, card: str) -> Path | None:
    """Localiza el Excel de referencia manual preferido sobre archivos generados."""
    candidates = [
        path
        for path in folder.glob("*.xlsx")
        if "Formato de Legalizaci" in path.name and card in path.name and "Plantilla" not in path.name
    ]
    if not candidates:
        return None

    manual = [
        path
        for path in candidates
        if MANUAL_NAME_PATTERN.search(path.name) and not GENERATED_NAME_PATTERN.search(path.name)
    ]
    if manual:
        return sorted(manual)[0]

    non_generated = [path for path in candidates if not GENERATED_NAME_PATTERN.search(path.name)]
    if non_generated:
        return sorted(non_generated)[0]
    return sorted(candidates)[0]


def find_manual_reference_for_compare(folder: Path, card: str) -> Path | None:
    """Alias de ``find_reference_legalization`` para comparaciones con referencia manual."""
    return find_reference_legalization(folder, card)
