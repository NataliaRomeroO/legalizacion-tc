"""Detección automática de columnas en plantillas Excel heterogéneas.

Escanea fila 18 (headers) y labels de footer para ubicar USD/CLP/SOL/COPS/IVA/total,
Documento Soporte y columna Legalizado. Plantillas antiguas pueden no tener columna SOL.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.worksheet.worksheet import Worksheet

HEADER_ROW = 18
FOOTER_LABEL = "valor total"


@dataclass
class TemplateLayout:
    """Índices de columnas detectados en una plantilla Excel de legalización."""

    col_usd: int = 8
    col_clp: int = 9
    col_sol: int | None = None
    col_cops: int = 10
    col_iva: int = 11
    col_total: int | None = 12
    col_soporte: int | None = None
    col_legalizado: int | None = None
    extractos_row: int | None = None
    extractos_label_col: int | None = None
    extractos_value_col: int | None = None
    checkpoint_row: int | None = None
    checkpoint_label_col: int | None = None
    checkpoint_value_col: int | None = None
    sum_columns: list[int] = field(default_factory=list)


def _header_text(ws: Worksheet, col: int) -> str:
    """Texto del encabezado en fila 18, normalizado a mayúsculas."""
    value = ws.cell(row=HEADER_ROW, column=col).value
    return str(value).upper() if value is not None else ""


def detect_template_layout(ws: Worksheet) -> TemplateLayout:
    """Detecta columnas de moneda, soporte, legalizado y filas de footer en la hoja."""
    layout = TemplateLayout()
    has_sol = False
    for col in range(1, 16):
        header = _header_text(ws, col)
        if not header:
            continue
        if "USD" in header:
            layout.col_usd = col
        elif "SOL" in header:
            layout.col_sol = col
            has_sol = True
        elif "COPS" in header:
            layout.col_cops = col
        elif "CLP" in header:
            layout.col_clp = col
        elif "IVA" in header:
            layout.col_iva = col
        elif "TOTAL" in header and "COMPRA" in header:
            layout.col_total = col
        elif "DOCUMENTO" in header and "SOPORTE" in header:
            layout.col_soporte = col

    sum_cols = {layout.col_usd, layout.col_clp, layout.col_cops, layout.col_iva}
    if has_sol and layout.col_sol is not None:
        sum_cols.add(layout.col_sol)
    if layout.col_total:
        sum_cols.add(layout.col_total)
    layout.sum_columns = sorted(sum_cols)

    for row_idx in range(HEADER_ROW, min(ws.max_row, 60) + 1):
        for col_idx in range(1, 16):
            value = ws.cell(row=row_idx, column=col_idx).value
            if not isinstance(value, str):
                continue
            lower = value.strip().lower()
            if "extractos o movimientos" in lower:
                layout.extractos_row = row_idx
                layout.extractos_label_col = col_idx
                layout.extractos_value_col = _value_column(ws, row_idx, col_idx)
            if lower == "checkpoint":
                layout.checkpoint_row = row_idx
                layout.checkpoint_label_col = col_idx
                layout.checkpoint_value_col = _value_column(ws, row_idx, col_idx)

    if layout.col_soporte:
        layout.col_legalizado = _detect_legalizado_column(ws, layout.col_soporte)
    else:
        last_header_col = max(
            (col for col in range(1, 16) if _header_text(ws, col)),
            default=layout.col_iva,
        )
        layout.col_legalizado = last_header_col + 1

    return layout


def _detect_legalizado_column(ws: Worksheet, after_col: int) -> int:
    """Ubica la primera columna vacía tras Documento Soporte para la etiqueta Legalizado."""
    for col_idx in range(after_col + 1, 16):
        header = ws.cell(row=HEADER_ROW, column=col_idx).value
        if header is None or str(header).strip() == "":
            return col_idx
    return after_col + 1


def _value_column(ws: Worksheet, row_idx: int, label_col: int) -> int:
    """Columna del valor numérico asociado a una etiqueta de footer (extractos/checkpoint)."""
    for col_idx in range(label_col + 1, 16):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None:
            continue
        if isinstance(value, str) and value.strip():
            continue
        return col_idx
    return label_col + 1


def total_sum_column(layout: TemplateLayout) -> int:
    """Columna preferida para sumar totales: TOTAL si existe, si no COPS."""
    return layout.col_total or layout.col_cops
