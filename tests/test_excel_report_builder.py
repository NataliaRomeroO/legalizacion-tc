"""Tests de generación/merge Excel: footer, fórmulas, highlight, append y GMF."""

from legalizacion_tc.config import load_settings
from legalizacion_tc.excel_report_builder import (
    HEADER_ROW,
    build_legalization_workbook,
    output_filename,
)
from legalizacion_tc.extract_parser import parse_extract
from legalizacion_tc.models import CardMetadata, ExtractData, LegalizationRow


def _realistic_template(path, tmp_path):
    """Helper de prueba: realistic template."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    headers = [
        "Nº Factura", "Nit Proveedor", "Razon social", "Detalle del Gasto",
        "Articulo Contable", "Centro de costo", "Moneda",
        "Valor base de la compra en USD", "Valor base de la compra en CLP",
        "Valor base de la compra en COPS", "Iva de la compa", "Valor total compra",
        "Documento Soporte",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    for row_idx in range(19, 24):
        ws.cell(row=row_idx, column=1, value=f"slot{row_idx}")
    fill = PatternFill("solid", fgColor="D9E1F2")
    border = Border(*(Side("thin"),) * 4)
    for col in range(7, 13):
        cell = ws.cell(row=24, column=col)
        cell.fill = fill
        cell.border = border
    ws.cell(row=24, column=7, value="Valor total")
    ws.cell(row=27, column=1, value="Autorizado por:")
    ws.cell(row=27, column=9, value="Extractos o movimientos")
    ws.cell(row=28, column=9, value="Checkpoint")
    ws.cell(row=30, column=1, value="Tramitado por:")
    ws.cell(row=33, column=1, value="Firma caja:")
    ws.cell(row=34, column=2, value="Firma")
    tpl = tmp_path / path
    wb.save(tpl)
    wb.close()
    return tpl


def _sample_rows(count: int) -> list[LegalizationRow]:
    """Helper de prueba: sample rows."""
    return [
        LegalizationRow(
            numero_factura=f"F{i}",
            nit_proveedor="900123456",
            razon_social="Proveedor SAS",
            detalle_gasto=f"TC 3333 GASTO {i}",
            articulo_contable="5195200005",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_usd=0.0,
            valor_base_cops=100.0 * (i + 1),
            iva=0.0,
            valor_total_compra_cop=100.0 * (i + 1),
            documento_soporte="SI",
        )
        for i in range(count)
    ]


def _footer_row(ws) -> int | None:
    """Helper de prueba: footer row."""
    for row_idx in range(19, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=7).value
        if isinstance(label, str) and label.strip().lower() == "valor total":
            return row_idx
    return None


def _find_row_with(ws, col: int, text: str) -> int | None:
    """Helper de prueba: find row with."""
    for row_idx in range(19, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=col).value
        if isinstance(value, str) and text in value:
            return row_idx
    return None


def test_few_rows_preserves_footer_and_signatures(settings, tmp_path):
    """Verifica few rows preserves footer and signatures."""
    template = _realistic_template("tpl_few.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=600.0,
        source_filename="mov.xlsx",
    )
    out = tmp_path / "out_few.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(3),
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    assert _footer_row(ws) == 24
    assert ws.cell(row=24, column=10).border.left.style == "thin"
    assert _find_row_with(ws, 1, "Firma caja:") == 33
    assert ws.cell(row=27, column=9).value == "Extractos o movimientos"
    assert ws.cell(row=27, column=10).value == 600.0
    assert ws.cell(row=28, column=10).value == "=J27-L24"
    assert ws.cell(row=22, column=1).value is None
    assert ws.cell(row=23, column=1).value is None
    wb.close()


def test_many_rows_shifts_footer_and_signatures(settings, tmp_path):
    """Verifica many rows shifts footer and signatures."""
    template = _realistic_template("tpl_many.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=3600.0,
        source_filename="mov.xlsx",
    )
    out = tmp_path / "out_many.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(8),
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    footer = _footer_row(ws)
    assert footer == 27
    firma_row = _find_row_with(ws, 1, "Firma caja:")
    assert firma_row == 36
    extractos_row = _find_row_with(ws, 9, "Extractos o movimientos")
    assert extractos_row == 30
    checkpoint_row = _find_row_with(ws, 9, "Checkpoint")
    assert checkpoint_row == 31
    assert ws.cell(row=checkpoint_row, column=10).value == f"=J{extractos_row}-L{footer}"
    assert ws.cell(row=footer + 1, column=9).value is None
    assert ws.cell(row=footer + 2, column=9).value is None
    wb.close()


def _count_extractos_labels(ws, footer: int) -> int:
    """Helper de prueba: count extractos labels."""
    count = 0
    for row_idx in range(footer + 1, ws.max_row + 1):
        for col in (7, 9):
            value = ws.cell(row=row_idx, column=col).value
            if isinstance(value, str) and "Extractos o movimientos" in value:
                count += 1
    return count


def test_append_clears_stale_extractos_block(settings, tmp_path):
    """Regresión Excel 1111: footer en 26, I27:J28 vacíos, bloque activo en 29-30."""
    template = _realistic_template("tpl_stale_footer.xlsx", tmp_path)
    extract = ExtractData(
        card="1111",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=134318.63,
        source_filename="mov.xlsx",
    )
    out = tmp_path / "out_stale_footer.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("1111", "Demo User A", "100-Demo"),
        _sample_rows(7),
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    footer = _footer_row(ws)
    assert footer == 26
    assert ws.cell(row=27, column=9).value is None
    assert ws.cell(row=27, column=10).value is None
    assert ws.cell(row=28, column=9).value is None
    assert ws.cell(row=28, column=10).value is None
    assert ws.cell(row=29, column=9).value == "Extractos o movimientos"
    assert ws.cell(row=29, column=10).value == 134318.63
    assert ws.cell(row=30, column=9).value == "Checkpoint"
    assert ws.cell(row=30, column=10).value == "=J29-L26"
    wb.close()


def test_no_duplicate_extractos_labels(settings, tmp_path):
    """Verifica no duplicate extractos labels."""
    template = _realistic_template("tpl_no_dup.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=3600.0,
        source_filename="mov.xlsx",
    )
    out = tmp_path / "out_no_dup.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(8),
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    footer = _footer_row(ws)
    assert _count_extractos_labels(ws, footer) == 1
    wb.close()


def test_excel_footer_formulas(fixture_dir, card_meta, tmp_path, settings):
    """Verifica excel footer formulas."""
    extract_path = fixture_dir / "Mov TC 1111 Corte 27 de Mayo.xlsx"
    template = fixture_dir / "plantilla_base.xlsx"
    extract = parse_extract(extract_path)
    rows = [
        LegalizationRow(
            numero_factura="FAC-700001",
            nit_proveedor="IE 1234567XX",
            razon_social="PROVEEDOR WEB INC",
            detalle_gasto="TC 1111 SERVICIO APLICACIONES",
            articulo_contable="5135950002 - SERVICIOS APLICACIONES",
            centro_costo="100-Demo",
            moneda="USD",
            valor_base_usd=15.0,
            valor_base_cops=56000.0,
            iva=0.0,
            valor_total_compra_cop=56000.0,
            documento_soporte="SI",
        )
    ]
    out = tmp_path / output_filename(extract)
    build_legalization_workbook(settings, template, out, extract, card_meta, rows)
    from openpyxl import load_workbook

    from legalizacion_tc.template_layout import detect_template_layout

    wb = load_workbook(out)
    ws = wb.active
    layout = detect_template_layout(ws)
    total_col = layout.col_total or layout.col_cops
    assert ws["B13"].value.startswith("LEGALIZACIÓN TC 1111")
    assert ws.cell(row=19, column=total_col).value == 56000.0
    footer_row = None
    for row_idx in range(19, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=7).value
        if isinstance(label, str) and label.strip().lower() == "valor total":
            footer_row = row_idx
            break
    assert footer_row is not None
    assert "SUM" in str(ws.cell(row=footer_row, column=total_col).value)
    wb.close()


def test_iva_cell_clears_stale_formula(settings, tmp_path):
    """Verifica iva cell clears stale formula."""
    from openpyxl import Workbook, load_workbook

    from legalizacion_tc.excel_report_builder import HEADER_ROW, build_legalization_workbook
    from legalizacion_tc.template_layout import detect_template_layout

    tpl = tmp_path / "tpl.xlsx"
    wb = Workbook()
    ws = wb.active
    headers = [
        "Nº Factura", "Nit Proveedor", "Razon social", "Detalle del Gasto",
        "Articulo Contable", "Centro de costo", "Moneda",
        "Valor base de la compra en USD", "Valor base de la compra en CLP",
        "Valor base de la compra en COPS", "Iva de la compa", "Valor total compra",
        "Documento Soporte",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    ws.cell(row=19, column=11, value="=K20")
    ws.cell(row=22, column=7, value="Valor total")
    wb.save(tpl)
    wb.close()

    extract = ExtractData(
        card="4444",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=100.0,
        source_filename="mov.xlsx",
    )
    rows = [
        LegalizationRow(
            numero_factura="REST64319",
            nit_proveedor="9007777777",
            razon_social="SPINNING CENTER GYM SAS",
            detalle_gasto="TC 4444 CONSUMO BAR",
            articulo_contable="",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_cops=37694.0,
            iva=0.0,
            valor_total_compra_cop=37694.0,
            documento_soporte="SI",
        ),
        LegalizationRow(
            numero_factura="COFE3377962",
            nit_proveedor="9006666666",
            razon_social="PROVEEDOR MAYORISTA S.A.S.",
            detalle_gasto="TC 4444 COMPRA MERCADO",
            articulo_contable="",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_cops=2327035.0,
            iva=426861.0,
            valor_total_compra_cop=2753900.0,
            documento_soporte="SI",
        ),
    ]
    out = tmp_path / "out.xlsx"
    build_legalization_workbook(settings, tpl, out, extract, CardMetadata("4444", "Test", "059"), rows)
    wb = load_workbook(out)
    ws = wb.active
    layout = detect_template_layout(ws)
    assert ws.cell(row=19, column=layout.col_iva).data_type != "f"
    assert ws.cell(row=19, column=layout.col_iva).value is None
    assert ws.cell(row=20, column=layout.col_iva).value == 426861.0
    wb.close()


def test_unmatched_row_highlighted_yellow(settings, tmp_path):
    """Verifica unmatched row highlighted yellow."""
    from openpyxl import load_workbook

    template = _realistic_template("tpl_review.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=85000.0,
        source_filename="mov.xlsx",
    )
    rows = [
        LegalizationRow(
            numero_factura="",
            nit_proveedor="",
            razon_social="",
            detalle_gasto="TC 3333 RESTAURANTE SIN FACTURA",
            articulo_contable="",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_usd=0.0,
            valor_base_cops=85000.0,
            iva=0.0,
            valor_total_compra_cop=85000.0,
            documento_soporte="",
            needs_review=True,
        )
    ]
    out = tmp_path / "out_review.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        rows,
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=19, column=1).fill.fgColor.rgb in ("FFFF00", "00FFFF00")
    assert ws.cell(row=19, column=13).value is None
    wb.close()


def test_suggested_row_with_nit_still_highlighted_yellow(settings, tmp_path):
    """Verifica suggested row with nit still highlighted yellow."""
    from openpyxl import load_workbook

    template = _realistic_template("tpl_suggested.xlsx", tmp_path)
    extract = ExtractData(
        card="5555",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=180000.0,
        source_filename="mov.xlsx",
    )
    rows = [
        LegalizationRow(
            numero_factura="B050-DEMO001",
            nit_proveedor="20987654321",
            razon_social="RESTAURANTE DEMO S.A.S",
            detalle_gasto="TC 5555 GASTO REST DEMO 24 DE MAYO",
            articulo_contable="",
            centro_costo="100-Demo",
            moneda="SOL",
            valor_base_sol=129.0,
            valor_base_cops=180000.0,
            iva=0.0,
            valor_total_compra_cop=180000.0,
            documento_soporte="NO",
            needs_review=True,
        )
    ]
    out = tmp_path / "out_suggested.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("5555", "Horacio", "100-Demo"),
        rows,
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=19, column=2).value == "20987654321"
    assert ws.cell(row=19, column=1).fill.fgColor.rgb in ("FFFF00", "00FFFF00")
    wb.close()


def test_documento_soporte_written_to_excel(settings, tmp_path):
    """Verifica documento soporte written to excel."""
    from openpyxl import load_workbook

    template = _realistic_template("tpl_soporte.xlsx", tmp_path)
    extract = ExtractData(
        card="1111",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=150000.0,
        source_filename="mov.xlsx",
    )
    rows = [
        LegalizationRow(
            numero_factura="INV-1",
            nit_proveedor="900123456",
            razon_social="AWS",
            detalle_gasto="TC 1111 GASTO AWS",
            articulo_contable="5135950002",
            centro_costo="100-Demo",
            moneda="USD",
            valor_base_usd=100.0,
            valor_base_cops=400000.0,
            iva=0.0,
            valor_total_compra_cop=400000.0,
            documento_soporte="SI",
        ),
        LegalizationRow(
            numero_factura="FAC-COP",
            nit_proveedor="900999999",
            razon_social="LOCAL SAS",
            detalle_gasto="TC 1111 GASTO LOCAL",
            articulo_contable="5195200005",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_usd=0.0,
            valor_base_cops=50000.0,
            iva=0.0,
            valor_total_compra_cop=50000.0,
            documento_soporte="NO",
        ),
    ]
    out = tmp_path / "out_soporte.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("1111", "Test User", "100-Demo"),
        rows,
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=19, column=13).value == "SI"
    assert ws.cell(row=20, column=13).value == "NO"
    wb.close()


def _full_template_with_sol(path, tmp_path):
    """Plantilla completa: USD, CLP, SOL, COPS, IVA, Total, Soporte."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    headers = [
        "Nº Factura", "Nit Proveedor", "Razon social", "Detalle del Gasto",
        "Articulo Contable", "Centro de costo", "Moneda",
        "Valor base de la compra en USD", "Valor base de la compra en CLP",
        "Valor base de la compra en SOL", "Valor base de la compra en COPS",
        "Iva de la compa", "Valor total compra", "Documento Soporte",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    for row_idx in range(19, 24):
        ws.cell(row=row_idx, column=1, value=f"slot{row_idx}")
    fill = PatternFill("solid", fgColor="D9E1F2")
    border = Border(*(Side("thin"),) * 4)
    for col in range(7, 14):
        cell = ws.cell(row=24, column=col)
        cell.fill = fill
        cell.border = border
    ws.cell(row=24, column=7, value="Valor total")
    tpl = tmp_path / path
    wb.save(tpl)
    wb.close()
    return tpl


def test_foreign_currency_columns_in_excel(settings, tmp_path):
    """Verifica foreign currency columns in excel."""
    from openpyxl import load_workbook

    template = _full_template_with_sol("tpl_fx.xlsx", tmp_path)
    extract = ExtractData(
        card="1111",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=400000.0,
        source_filename="mov.xlsx",
    )
    rows = [
        LegalizationRow(
            numero_factura="INV-USD",
            nit_proveedor="900123456",
            razon_social="AWS",
            detalle_gasto="TC 1111 GASTO AWS",
            articulo_contable="5135950002",
            centro_costo="100-Demo",
            moneda="USD",
            valor_base_usd=100.0,
            valor_base_cops=400000.0,
            iva=0.0,
            valor_total_compra_cop=400000.0,
            documento_soporte="SI",
        ),
        LegalizationRow(
            numero_factura="FAC-COP",
            nit_proveedor="900999999",
            razon_social="LOCAL SAS",
            detalle_gasto="TC 1111 GASTO LOCAL",
            articulo_contable="5195200005",
            centro_costo="100-Demo",
            moneda="COP",
            valor_base_usd=0.0,
            valor_base_cops=50000.0,
            iva=9500.0,
            valor_total_compra_cop=59500.0,
            documento_soporte="NO",
        ),
    ]
    out = tmp_path / "out_fx.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("1111", "Test User", "100-Demo"),
        rows,
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=19, column=7).value == "USD"
    assert ws.cell(row=19, column=8).value == 100.0
    assert ws.cell(row=19, column=9).value is None
    assert ws.cell(row=19, column=10).value is None
    assert ws.cell(row=19, column=11).value == 400000.0
    assert ws.cell(row=19, column=12).value is None
    assert ws.cell(row=19, column=13).value == 400000.0
    assert ws.cell(row=20, column=7).value == "COP"
    assert ws.cell(row=20, column=11).value == 50000.0
    assert ws.cell(row=20, column=12).value == 9500.0
    assert ws.cell(row=20, column=13).value == 59500.0
    wb.close()


def test_merge_preserves_two_gmf_rows_and_appends_new(settings, tmp_path, monkeypatch):
    """Verifica merge preserves two gmf rows and appends new."""
    from openpyxl import load_workbook

    from legalizacion_tc import excel_report_builder
    from legalizacion_tc.excel_report_builder import (
        DATA_START_ROW,
        build_legalization_workbook,
        merge_legalization_workbook,
    )

    monkeypatch.setattr(
        excel_report_builder,
        "execution_period_month",
        lambda **kwargs: "JULIO",
    )

    template = _realistic_template("tpl_gmf_base.xlsx", tmp_path)
    extract = ExtractData(
        card="1111",
        period_month="MAYO",
        period_year=2026,
        transactions=[],
        total_cop=60000.0,
        source_filename="mov.xlsx",
    )
    base = tmp_path / "base_mayo.xlsx"
    build_legalization_workbook(
        settings,
        template,
        base,
        extract,
        CardMetadata("1111", "Gabriel", "100-Demo"),
        [
            LegalizationRow(
                numero_factura="FAC-700001",
                nit_proveedor="IE 1234567XX",
                razon_social="PROVEEDOR WEB",
                detalle_gasto="TC 1111 SERVICIO APLICACIONES",
                articulo_contable="5135950002",
                centro_costo="100-Demo",
                moneda="USD",
                valor_base_usd=15.0,
                valor_base_cops=56000.0,
                iva=0.0,
                valor_total_compra_cop=56000.0,
                documento_soporte="SI",
            ),
            LegalizationRow(
                numero_factura="",
                nit_proveedor="8600000001",
                razon_social="BANCOLOMBIA ",
                detalle_gasto="TC 1111 GMF ",
                articulo_contable="5115950001 - GMF IMPUESTOS",
                centro_costo="100-Demo",
                moneda="COP",
                valor_base_cops=225.99,
                iva=0.0,
                valor_total_compra_cop=225.99,
                documento_soporte="NO",
            ),
            LegalizationRow(
                numero_factura="",
                nit_proveedor="8600000001",
                razon_social="BANCOLOMBIA ",
                detalle_gasto="TC 1111 GMF ",
                articulo_contable="5115950001 - GMF IMPUESTOS",
                centro_costo="100-Demo",
                moneda="COP",
                valor_base_cops=309.14,
                iva=0.0,
                valor_total_compra_cop=309.14,
                documento_soporte="NO",
            ),
        ],
    )

    out = tmp_path / "Formato de Legalización TC 1111 - 3 - JULIO - 2026.xlsx"
    merge_legalization_workbook(
        settings,
        base,
        out,
        extract,
        CardMetadata("1111", "Gabriel", "100-Demo"),
        [
            LegalizationRow(
                numero_factura="94964403-0024",
                nit_proveedor="511111111111-99",
                razon_social="PROVEEDOR IA",
                detalle_gasto="TC 1111 SUSCRIPCION APP IA PRO 28 DE MAYO",
                articulo_contable="5135950002",
                centro_costo="100-Demo",
                moneda="USD",
                valor_base_usd=20.0,
                valor_base_cops=75000.0,
                iva=0.0,
                valor_total_compra_cop=75000.0,
                documento_soporte="SI",
            ),
        ],
    )

    wb = load_workbook(out)
    ws = wb.active
    gmf_rows = [
        row
        for row in range(DATA_START_ROW, DATA_START_ROW + 10)
        if ws.cell(row=row, column=4).value == "TC 1111 GMF "
    ]
    assert len(gmf_rows) == 2
    assert ws.cell(row=gmf_rows[0], column=12).value == 225.99
    assert ws.cell(row=gmf_rows[1], column=12).value == 309.14
    assert ws.cell(row=DATA_START_ROW, column=1).value == "FAC-700001"
    assert ws.cell(row=DATA_START_ROW + 3, column=1).value == "94964403-0024"
    assert not any(
        ws.cell(row=row, column=12).value == 535.13
        for row in range(DATA_START_ROW, DATA_START_ROW + 10)
    )
    wb.close()


def test_append_preserves_existing_rows_and_relabels(settings, tmp_path, monkeypatch):
    """Verifica append preserves existing rows and relabels."""
    from openpyxl import load_workbook

    from legalizacion_tc import excel_report_builder
    from legalizacion_tc.excel_report_builder import merge_legalization_workbook

    monkeypatch.setattr(
        excel_report_builder,
        "execution_period_month",
        lambda **kwargs: "JUNIO",
    )

    template = _realistic_template("tpl_append.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="JUNIO",
        period_year=2026,
        transactions=[],
        total_cop=900.0,
        source_filename="mov.xlsx",
    )
    first_out = tmp_path / "first.xlsx"
    build_legalization_workbook(
        settings,
        template,
        first_out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(2),
    )

    append_out = tmp_path / "append.xlsx"
    merge_legalization_workbook(
        settings,
        first_out,
        append_out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(1),
    )

    wb = load_workbook(append_out)
    ws = wb.active
    assert ws.cell(row=19, column=1).value == "F0"
    assert ws.cell(row=20, column=1).value == "F1"
    assert ws.cell(row=21, column=1).value == "F0"
    assert ws.cell(row=19, column=14).value == "Legalizado en junio corte 1"
    assert ws.cell(row=20, column=14).value == "Legalizado en junio corte 1"
    assert ws.cell(row=21, column=14).value == "Legalizado en junio"
    wb.close()


def test_append_uses_execution_month_not_extract_period(settings, tmp_path, monkeypatch):
    """Verifica append uses execution month not extract period."""
    from openpyxl import load_workbook

    from legalizacion_tc import excel_report_builder
    from legalizacion_tc.excel_report_builder import merge_legalization_workbook

    monkeypatch.setattr(
        excel_report_builder,
        "execution_period_month",
        lambda **kwargs: "JULIO",
    )

    template = _realistic_template("tpl_cross_month.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="JUNIO",
        period_year=2026,
        transactions=[],
        total_cop=900.0,
        source_filename="mov.xlsx",
    )
    first_out = tmp_path / "first_cross.xlsx"
    build_legalization_workbook(
        settings,
        template,
        first_out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(1),
    )
    wb = load_workbook(first_out)
    ws = wb.active
    ws.cell(row=19, column=14, value="Legalizado en junio")
    wb.save(first_out)
    wb.close()

    append_out = tmp_path / "append_cross.xlsx"
    merge_legalization_workbook(
        settings,
        first_out,
        append_out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(1),
    )

    wb = load_workbook(append_out)
    ws = wb.active
    assert ws.cell(row=19, column=14).value == "Legalizado en junio"
    assert ws.cell(row=20, column=14).value == "Legalizado en julio"
    wb.close()


def test_create_writes_batch_label(settings, tmp_path, monkeypatch):
    """Verifica create writes batch label."""
    from openpyxl import load_workbook

    from legalizacion_tc import excel_report_builder

    monkeypatch.setattr(
        excel_report_builder,
        "execution_period_month",
        lambda **kwargs: "JUNIO",
    )

    template = _realistic_template("tpl_label.xlsx", tmp_path)
    extract = ExtractData(
        card="3333",
        period_month="JUNIO",
        period_year=2026,
        transactions=[],
        total_cop=300.0,
        source_filename="mov.xlsx",
    )
    out = tmp_path / "labeled.xlsx"
    build_legalization_workbook(
        settings,
        template,
        out,
        extract,
        CardMetadata("3333", "Test User", "100-Demo"),
        _sample_rows(1),
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=19, column=14).value == "Legalizado en junio"
    wb.close()
