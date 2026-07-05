"""Tests de deduplicación en re-ejecución/append (``legalization_dedup``)."""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from legalizacion_tc.date_normalizer import parse_detalle_date_suffix
from legalizacion_tc.excel_report_builder import DATA_START_ROW, create_minimal_template
from legalizacion_tc.legalization_dedup import (
    LegalizedState,
    filter_matches_for_append,
    is_match_already_legalized,
    is_transaction_legalized,
    legalized_state_from_paths,
    legalized_state_from_workbook,
    merge_legalized_states,
    normalize_desc,
    transaction_fingerprint,
)
from legalizacion_tc.models import InvoiceData, MatchResult, Transaction


def _tx(amount: float, desc: str, tx_date: date | None = None) -> Transaction:
    """Helper de prueba: tx."""
    return Transaction(
        card="2222",
        tx_date=tx_date or date(2026, 6, 5),
        description=desc,
        currency="COP",
        amount_cop=amount,
    )


def test_parse_detalle_date_suffix_spanish():
    """Verifica parse detalle date suffix spanish."""
    assert parse_detalle_date_suffix("TC 1111 PROVEEDOR WEB 9 DE MAYO", 2026) == date(
        2026, 5, 9
    )
    assert parse_detalle_date_suffix("28 DE MAYO", 2026) == date(2026, 5, 28)


def test_transaction_fingerprint():
    """Verifica transaction fingerprint."""
    tx = _tx(150000.0, "DELIVERY  *REST")
    assert transaction_fingerprint(tx) == (
        "2026-06-05",
        150000.0,
        normalize_desc("DELIVERY  *REST"),
    )


def test_filter_skips_already_legalized():
    """Verifica filter skips already legalized."""
    tx = _tx(150000.0, "DELIVERY  *REST")
    state = LegalizedState(
        fingerprints={transaction_fingerprint(tx)},
        date_amount_keys={("2026-06-05", 150000.0)},
    )
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="UNMATCHED",
        documento_soporte="NO",
    )
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_filter_skips_gmf_when_same_fingerprint():
    """Verifica filter skips gmf when same fingerprint."""
    tx = _tx(1200.0, "GMF", date(2026, 6, 1))
    tx.is_gmf = True
    state = LegalizedState(fingerprints={transaction_fingerprint(tx)})
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="GMF",
        documento_soporte="NO",
    )
    assert is_transaction_legalized(tx, state)
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_filter_skips_gmf_by_amount_from_excel_row(tmp_path: Path):
    """Verifica filter skips gmf by amount from excel row."""
    path = tmp_path / "legalization.xlsx"
    create_minimal_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B13"] = "LEGALIZACIÓN TC 1111 - MAYO 2026"
    ws.cell(row=DATA_START_ROW, column=4, value="TC 1111 GMF ")
    ws.cell(row=DATA_START_ROW, column=11, value=31446.47)
    wb.save(path)
    wb.close()

    state = legalized_state_from_workbook(path, default_year=2026)
    assert 31446.47 in state.gmf_amount_keys

    tx = _tx(31446.47, "GMF JURIDICO", date(2026, 5, 31))
    tx.is_gmf = True
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="GMF",
        documento_soporte="NO",
    )
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_filter_allows_new_gmf_amount_on_append():
    """Verifica filter allows new gmf amount on append."""
    state = LegalizedState(gmf_amount_keys={31446.47})
    tx = _tx(5000.0, "GMF", date(2026, 6, 30))
    tx.is_gmf = True
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="GMF",
        documento_soporte="NO",
    )
    kept, skipped = filter_matches_for_append([match], state)
    assert len(kept) == 1
    assert skipped == 0


def test_filter_skips_consolidated_gmf_equal_to_prior_sum():
    """PDF may report one GMF line equal to the sum of prior GMF rows."""
    state = LegalizedState(gmf_amount_keys={225.99, 309.14})
    tx = _tx(535.13, "GMF JURIDICO", date(2026, 5, 31))
    tx.is_gmf = True
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="GMF",
        documento_soporte="NO",
    )
    assert is_match_already_legalized(match, state)
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_filter_skips_ok_match_when_invoice_nit_exists():
    """Verifica filter skips ok match when invoice nit exists."""
    invoice = InvoiceData(
        source_filename="sq.pdf",
        numero_factura="FAC-700001",
        nit_proveedor="900123456",
        moneda="USD",
        valor_total_documento=20.0,
        legible=True,
    )
    tx = _tx(75000.0, "WEBAPP* DOMAIN", date(2026, 5, 9))
    state = LegalizedState(invoice_nit_keys={("FAC-700001", "900123456")})
    match = MatchResult(
        transaction=tx,
        invoice=invoice,
        status="OK",
        documento_soporte="SI",
    )
    assert is_match_already_legalized(match, state)
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_state_from_workbook_parses_spanish_detalle_and_invoice_nit(tmp_path: Path):
    """Verifica state from workbook parses spanish detalle and invoice nit."""
    path = tmp_path / "legalization.xlsx"
    create_minimal_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B13"] = "LEGALIZACIÓN TC 1111 - MAYO 2026"
    ws.cell(row=DATA_START_ROW, column=1, value="FAC-700001")
    ws.cell(row=DATA_START_ROW, column=2, value="900123456")
    ws.cell(
        row=DATA_START_ROW,
        column=4,
        value="TC 1111 SUSCRIPCION WEB 9 DE MAYO",
    )
    ws.cell(row=DATA_START_ROW, column=11, value=75000.0)
    wb.save(path)
    wb.close()

    state = legalized_state_from_workbook(path, default_year=2026)
    assert ("FAC-700001", "900123456") in state.invoice_nit_keys
    assert ("2026-05-09", 75000.0, "SUSCRIPCION WEB") in state.fingerprints


def test_state_from_workbook_normalizes_excel_float_invoice_and_nit(tmp_path: Path):
    """Manual Formatos often store invoice/NIT as Excel floats (700001.0, 53156.0)."""
    path = tmp_path / "legalization.xlsx"
    create_minimal_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B13"] = "LEGALIZACIÓN TC 1111 - MAYO 2026"
    # Minimal template has only one data row before footer; make room for two.
    ws.insert_rows(DATA_START_ROW + 1)
    # Proveedor web: numeric invoice + foreign NIT string
    ws.cell(row=DATA_START_ROW, column=1, value=700001.0)
    ws.cell(row=DATA_START_ROW, column=2, value="IE 1234567XX")
    ws.cell(row=DATA_START_ROW, column=4, value="TC 1111 SERVICIO APLICACIONES")
    ws.cell(row=DATA_START_ROW, column=11, value=56000.0)
    # Recibo genérico: text invoice + numeric NIT float
    ws.cell(row=DATA_START_ROW + 1, column=1, value="RECIBO DE CAJA 813")
    ws.cell(row=DATA_START_ROW + 1, column=2, value=53156.0)
    ws.cell(row=DATA_START_ROW + 1, column=4, value="TC 1111 _SERVICIO APLICACIONES")
    ws.cell(row=DATA_START_ROW + 1, column=11, value=3700.0)
    wb.save(path)
    wb.close()

    state = legalized_state_from_workbook(path, default_year=2026)
    assert ("700001", "1234567") in state.invoice_nit_keys
    assert ("RECIBO DE CAJA 813", "53156") in state.invoice_nit_keys
    # Must not keep float-string artifacts
    assert ("700001.0", "1234567") not in state.invoice_nit_keys
    assert ("RECIBO DE CAJA 813", "531560") not in state.invoice_nit_keys

    # Foreign invoice JSON often has null NIT; Excel may still have historico NIT.
    web_invoice = InvoiceData(
        source_filename="sq.pdf",
        numero_factura="700001",
        nit_proveedor=None,
        moneda="USD",
        valor_total_documento=15.0,
        legible=True,
    )
    recibo_invoice = InvoiceData(
        source_filename="db.pdf",
        numero_factura="RECIBO DE CAJA 813",
        nit_proveedor="53156",
        moneda="USD",
        valor_total_documento=0.9,
        legible=True,
    )
    matches = [
        MatchResult(
            transaction=_tx(56000.0, "WEBAPP* DOMAIN", date(2026, 5, 9)),
            invoice=web_invoice,
            status="OK",
            documento_soporte="SI",
        ),
        MatchResult(
            transaction=_tx(3700.0, "SERVICIO ARCHIVO", date(2026, 6, 5)),
            invoice=recibo_invoice,
            status="OK",
            documento_soporte="NO",
        ),
    ]
    kept, skipped = filter_matches_for_append(matches, state)
    assert kept == []
    assert skipped == 2


def test_state_from_workbook_skips_duplicate_on_reappend(tmp_path: Path):
    """Verifica state from workbook skips duplicate on reappend."""
    path = tmp_path / "legalization.xlsx"
    create_minimal_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B13"] = "LEGALIZACIÓN TC 1111 - MAYO 2026"
    for offset, amount in ((0, 100000.0), (1, 50000.0)):
        row = DATA_START_ROW + offset
        ws.cell(row=row, column=1, value="INV-001")
        ws.cell(row=row, column=2, value="900111222")
        ws.cell(row=row, column=4, value=f"TC 1111 COMPRA IVA SPLIT {10 + offset} DE MAYO")
        ws.cell(row=row, column=11, value=amount)
    wb.save(path)
    wb.close()

    state = legalized_state_from_workbook(path, default_year=2026)
    invoice = InvoiceData(
        source_filename="inv.pdf",
        numero_factura="INV-001",
        nit_proveedor="900111222",
        moneda="COP",
        valor_total_documento=150000.0,
        legible=True,
    )
    tx = _tx(150000.0, "COMPRA IVA", date(2026, 5, 10))
    match = MatchResult(
        transaction=tx,
        invoice=invoice,
        status="OK",
        documento_soporte="NO",
    )
    kept, skipped = filter_matches_for_append([match], state)
    assert kept == []
    assert skipped == 1


def test_merge_legalized_states_unions_keys():
    """Verifica merge legalized states unions keys."""
    a = LegalizedState(
        fingerprints={("2026-05-09", 100.0, "A")},
        invoice_nit_keys={("INV-1", "9001")},
    )
    b = LegalizedState(
        fingerprints={("2026-05-10", 200.0, "B")},
        gmf_amount_keys={50.0},
    )
    merged = merge_legalized_states(a, b)
    assert ("2026-05-09", 100.0, "A") in merged.fingerprints
    assert ("2026-05-10", 200.0, "B") in merged.fingerprints
    assert ("INV-1", "9001") in merged.invoice_nit_keys
    assert 50.0 in merged.gmf_amount_keys


def test_legalized_state_from_paths_merges_workbooks(tmp_path: Path):
    """Verifica legalized state from paths merges workbooks."""
    path_a = tmp_path / "a.xlsx"
    path_b = tmp_path / "b.xlsx"
    for path, numero, amount, day in (
        (path_a, "INV-A", 100000.0, 9),
        (path_b, "INV-B", 50000.0, 10),
    ):
        create_minimal_template(path)
        wb = load_workbook(path)
        ws = wb.active
        ws.cell(row=DATA_START_ROW, column=1, value=numero)
        ws.cell(row=DATA_START_ROW, column=2, value="900123456")
        ws.cell(
            row=DATA_START_ROW,
            column=4,
            value=f"TC 1111 COMPRA {day} DE MAYO",
        )
        ws.cell(row=DATA_START_ROW, column=11, value=amount)
        wb.save(path)
        wb.close()

    state = legalized_state_from_paths([path_a, path_b], default_year=2026)
    assert ("INV-A", "900123456") in state.invoice_nit_keys
    assert ("INV-B", "900123456") in state.invoice_nit_keys

    invoice_a = InvoiceData(
        source_filename="a.pdf",
        numero_factura="INV-A",
        nit_proveedor="900123456",
        moneda="COP",
        valor_total_documento=100000.0,
        legible=True,
    )
    match_a = MatchResult(
        transaction=_tx(100000.0, "COMPRA", date(2026, 5, 9)),
        invoice=invoice_a,
        status="OK",
        documento_soporte="NO",
    )
    invoice_new = InvoiceData(
        source_filename="c.pdf",
        numero_factura="INV-C",
        nit_proveedor="900123456",
        moneda="COP",
        valor_total_documento=75000.0,
        legible=True,
    )
    match_new = MatchResult(
        transaction=_tx(75000.0, "NUEVO", date(2026, 5, 11)),
        invoice=invoice_new,
        status="OK",
        documento_soporte="NO",
    )
    kept, skipped = filter_matches_for_append([match_a, match_new], state)
    assert skipped == 1
    assert len(kept) == 1
    assert kept[0].invoice.numero_factura == "INV-C"
