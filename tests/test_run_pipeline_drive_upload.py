"""Tests de upload Drive: preliminar update in-place; Formato siempre nuevo con v2."""

from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import pytest

from legalizacion_tc.drive_manager import DriveFile
from legalizacion_tc.folder_resolver import CardFolderContext
from legalizacion_tc.models import CardMetadata
from legalizacion_tc.run_pipeline import run_pipeline_for_card
from tests.build_fixtures import build_minimal_extract

EXTRACT_FILE_ID = "extract-file-id-123"
FOLDER_ID = "folder-drive-abc"
PRIOR_LEGALIZATION_ID = "prior-legalization-id"


@pytest.fixture
def drive_card_setup(tmp_path, monkeypatch):
    """Fixture o helper: drive card setup."""
    extract_name = "Mov TC 1111 Corte Mayo.xlsx"
    extract_local = tmp_path / extract_name
    build_minimal_extract(
        extract_local,
        "1111",
        rows=[
            {
                "Tarjeta": "*1111",
                "Fecha": "2026-05-10",
                "Descripcion": "COMPRA EJEMPLO",
                "Moneda": "COP",
                "Valor": 100000.0,
            }
        ],
    )

    extract_file = DriveFile(
        file_id=EXTRACT_FILE_ID,
        name=extract_name,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    folder_files: list[DriveFile] = [extract_file]

    update_calls: list[tuple[str, Path]] = []
    upload_calls: list[tuple[Path, str, str]] = []
    out_dir = tmp_path / "output" / "1111"

    def fake_list_folder_files(folder_id: str) -> list[DriveFile]:
        """Fixture o helper: fake list folder files."""
        assert folder_id == FOLDER_ID
        return list(folder_files)

    def fake_download_to_cache(file_id: str, name: str, card: str | None = None) -> Path:
        """Fixture o helper: fake download to cache."""
        dest = tmp_path / "cache" / (card or "x") / "downloads" / name
        dest.parent.mkdir(parents=True, exist_ok=True)
        if file_id == EXTRACT_FILE_ID:
            shutil.copy(extract_local, dest)
        return dest

    def fake_update_file(file_id: str, path: Path) -> str:
        """Fixture o helper: fake update file."""
        update_calls.append((file_id, Path(path)))
        return file_id

    def fake_upload_file(path: Path, folder_id: str, name: str) -> str:
        """Fixture o helper: fake upload file."""
        upload_calls.append((Path(path), folder_id, name))
        return "uploaded-legalization-id"

    def fake_file_link(file_id: str) -> str:
        """Fixture o helper: fake file link."""
        return f"https://drive.google.com/file/d/{file_id}/view"

    def fake_create_minimal_template(path: Path) -> None:
        """Fixture o helper: fake create minimal template."""
        path.parent.mkdir(parents=True, exist_ok=True)
        from legalizacion_tc.excel_report_builder import create_minimal_template

        create_minimal_template(path)

    def fake_download_file(_file_id: str, dest: Path) -> None:
        """Fixture o helper: fake download file."""
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        from legalizacion_tc.excel_report_builder import create_minimal_template

        create_minimal_template(dest)

    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.list_folder_files", fake_list_folder_files
    )
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.download_to_cache", fake_download_to_cache
    )
    monkeypatch.setattr("legalizacion_tc.run_pipeline.download_file", fake_download_file)
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.get_card_metadata",
        lambda *_a, **_k: CardMetadata("1111", "Demo User A", "100-Demo"),
    )
    monkeypatch.setattr("legalizacion_tc.run_pipeline.load_historico", lambda *_a, **_k: {})
    monkeypatch.setattr("legalizacion_tc.run_pipeline.update_file", fake_update_file)
    monkeypatch.setattr("legalizacion_tc.run_pipeline.upload_file", fake_upload_file)
    monkeypatch.setattr("legalizacion_tc.run_pipeline.file_link", fake_file_link)
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.create_minimal_template", fake_create_minimal_template
    )
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.output_cache_dir",
        lambda card=None: out_dir,
    )
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.execution_date",
        lambda **kwargs: date(2026, 7, 2),
    )

    context = CardFolderContext(
        card="1111",
        folder_id=FOLDER_ID,
        local_path=None,
        display_name="1111",
    )

    return {
        "context": context,
        "extract_file": extract_file,
        "folder_files": folder_files,
        "update_calls": update_calls,
        "upload_calls": upload_calls,
        "out_dir": out_dir,
        "tmp_path": tmp_path,
    }


def _extract_update_calls(setup: dict) -> list[tuple[str, Path]]:
    """Helper de prueba: extract update calls."""
    return [c for c in setup["update_calls"] if c[0] == EXTRACT_FILE_ID]


def _preliminar_upload_calls(setup: dict) -> list[tuple[Path, str, str]]:
    """Helper de prueba: preliminar upload calls."""
    return [c for c in setup["upload_calls"] if c[2].startswith("Mov TC")]


def _legalization_upload_calls(setup: dict) -> list[tuple[Path, str, str]]:
    """Helper de prueba: legalization upload calls."""
    return [c for c in setup["upload_calls"] if c[2].startswith("Formato de Legalización")]


def test_drive_updates_preliminar_in_place(drive_card_setup):
    """Verifica drive updates preliminar in place."""
    setup = drive_card_setup
    result = run_pipeline_for_card(
        setup["context"],
        skip_invoice_check=True,
        dry_run=False,
    )

    extract_updates = _extract_update_calls(setup)
    assert len(extract_updates) == 1
    assert extract_updates[0][1].name == setup["extract_file"].name
    assert _preliminar_upload_calls(setup) == []

    assert result.extract_file_link == f"https://drive.google.com/file/d/{EXTRACT_FILE_ID}/view"
    assert result.extract_update_mode == "update"


def test_drive_uploads_new_legalization_with_execution_day_name(drive_card_setup):
    """Verifica drive uploads new legalization with execution day name."""
    setup = drive_card_setup
    result = run_pipeline_for_card(
        setup["context"],
        skip_invoice_check=True,
        dry_run=False,
    )

    legalization_uploads = _legalization_upload_calls(setup)
    assert len(legalization_uploads) == 1
    assert legalization_uploads[0][1] == FOLDER_ID
    assert (
        legalization_uploads[0][2]
        == "Formato de Legalización TC 1111 - 2 - JULIO - 2026.xlsx"
    )
    assert result.legalization_mode == "create"
    assert result.output_version == 1
    assert result.output_filename == legalization_uploads[0][2]
    assert not any(c[0] != EXTRACT_FILE_ID for c in setup["update_calls"])


def test_drive_uses_v2_when_same_day_name_exists(drive_card_setup):
    """Verifica drive uses v2 when same day name exists."""
    setup = drive_card_setup
    setup["folder_files"].append(
        DriveFile(
            file_id=PRIOR_LEGALIZATION_ID,
            name="Formato de Legalización TC 1111 - 2 - JULIO - 2026.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    result = run_pipeline_for_card(
        setup["context"],
        skip_invoice_check=True,
        dry_run=False,
    )

    legalization_uploads = _legalization_upload_calls(setup)
    assert len(legalization_uploads) == 1
    assert (
        legalization_uploads[0][2]
        == "Formato de Legalización TC 1111 - 2 - JULIO v2 - 2026.xlsx"
    )
    assert result.output_version == 2
    assert result.legalization_mode == "create"
    assert not any(c[0] == PRIOR_LEGALIZATION_ID for c in setup["update_calls"])


def test_drive_merge_preserves_base_gmf_rows(drive_card_setup, monkeypatch):
    """Verifica drive merge preserves base gmf rows."""
    from openpyxl import Workbook, load_workbook

    from legalizacion_tc.excel_report_builder import DATA_START_ROW, HEADER_ROW

    setup = drive_card_setup
    base_name = "Formato de Legalización TC 1111 - MAYO- 2026.xlsx"
    base_local = setup["tmp_path"] / base_name

    wb = Workbook()
    ws = wb.active
    headers = [
        "Nº Factura", "Nit Proveedor", "Razon social", "Detalle del Gasto",
        "Articulo Contable", "Centro de costo", "Moneda",
        "Valor base de la compra en USD", "Valor base de la compra en CLP",
        "Valor base de la compra en COPS", "Iva de la compa", "Valor total compra",
        "Documento Soporte",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col, value=header)
    ws.cell(row=DATA_START_ROW, column=1, value="FAC-700001")
    ws.cell(row=DATA_START_ROW, column=2, value="IE 1234567XX")
    ws.cell(row=DATA_START_ROW, column=4, value="TC 1111 SERVICIO APLICACIONES")
    ws.cell(row=DATA_START_ROW, column=7, value="USD")
    ws.cell(row=DATA_START_ROW, column=12, value=56000.0)
    ws.cell(row=DATA_START_ROW + 1, column=2, value="8600000001")
    ws.cell(row=DATA_START_ROW + 1, column=4, value="TC 1111 GMF ")
    ws.cell(row=DATA_START_ROW + 1, column=7, value="COP")
    ws.cell(row=DATA_START_ROW + 1, column=12, value=225.99)
    ws.cell(row=DATA_START_ROW + 2, column=2, value="8600000001")
    ws.cell(row=DATA_START_ROW + 2, column=4, value="TC 1111 GMF ")
    ws.cell(row=DATA_START_ROW + 2, column=7, value="COP")
    ws.cell(row=DATA_START_ROW + 2, column=12, value=309.14)
    ws.cell(row=DATA_START_ROW + 3, column=7, value="Valor total")
    wb.save(base_local)
    wb.close()

    setup["folder_files"].append(
        DriveFile(
            file_id=PRIOR_LEGALIZATION_ID,
            name=base_name,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            modified_time=1.0,
        )
    )

    def fake_download_file(file_id: str, dest: Path) -> None:
        """Fixture o helper: fake download file."""
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        if file_id == PRIOR_LEGALIZATION_ID:
            shutil.copy(base_local, dest)
            return
        from legalizacion_tc.excel_report_builder import create_minimal_template

        create_minimal_template(dest)

    monkeypatch.setattr("legalizacion_tc.run_pipeline.download_file", fake_download_file)

    result = run_pipeline_for_card(
        setup["context"],
        skip_invoice_check=True,
        dry_run=False,
    )

    legalization_uploads = _legalization_upload_calls(setup)
    assert len(legalization_uploads) == 1
    assert legalization_uploads[0][2].startswith(
        "Formato de Legalización TC 1111 - 2 - JULIO"
    )
    assert not any(c[0] == PRIOR_LEGALIZATION_ID for c in setup["update_calls"])

    out_path = Path(result.output_path)
    assert out_path.exists()
    wb = load_workbook(out_path)
    ws = wb.active
    gmf_rows = [
        row
        for row in range(DATA_START_ROW, DATA_START_ROW + 10)
        if ws.cell(row=row, column=4).value == "TC 1111 GMF "
    ]
    assert len(gmf_rows) == 2
    assert ws.cell(row=gmf_rows[0], column=12).value == 225.99
    assert ws.cell(row=gmf_rows[1], column=12).value == 309.14
    assert ws.cell(row=DATA_START_ROW, column=1).value == "FAC-700001"
    wb.close()


def test_drive_updates_preliminar_without_new_legalization_rows(drive_card_setup, monkeypatch):
    """Verifica drive updates preliminar without new legalization rows."""
    setup = drive_card_setup
    monkeypatch.setattr(
        "legalizacion_tc.run_pipeline.build_legalization_rows",
        lambda *_a, **_k: ([], []),
    )

    result = run_pipeline_for_card(
        setup["context"],
        skip_invoice_check=True,
        dry_run=False,
    )

    assert len(_extract_update_calls(setup)) == 1
    assert _preliminar_upload_calls(setup) == []
    assert _legalization_upload_calls(setup) == []
    assert result.extract_update_mode == "update"
    assert result.appended_row_count == 0
