"""Tests de flags Validación, observaciones, near-miss y update del preliminar Excel."""

from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from legalizacion_tc.config import Settings
from legalizacion_tc.extract_updater import (
    OBSERVATIONS_COLUMN,
    VALIDATION_COLUMN,
    apply_extract_review_columns,
)
from legalizacion_tc.models import ExtractData, InvoiceData, MatchResult, Transaction
from legalizacion_tc.reconciliation_engine import reconcile
from legalizacion_tc.validation_notes import (
    find_near_miss_candidate,
    format_ok_observation,
    pick_best_from_candidates,
    resolve_suggested_invoice,
    validation_flag,
)


def _rest_demo_invoice() -> InvoiceData:
    """Helper de prueba: sushi invoice."""
    return InvoiceData(
        source_filename="RESTAURANTE DEMO -F001.jpeg",
        numero_factura="B050-DEMO001",
        razon_social="RESTAURANTE DEMO S.A.S",
        moneda="SOL",
        valor_total_documento=129.0,
        fecha_factura=date(2026, 5, 24),
        detalle_gasto="TC 5555 GASTO REST DEMO 24 DE MAYO",
        legible=True,
    )


def test_validation_flag_ok_revisar_and_no():
    """Verifica validation flag ok revisar and no."""
    ok = MatchResult(
        transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
        invoice=None,
        status="OK",
        documento_soporte="SI",
    )
    assert validation_flag(ok) == "OK"
    ok_review = MatchResult(
        transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
        invoice=None,
        status="OK",
        documento_soporte="SI",
        needs_review=True,
    )
    assert validation_flag(ok_review) == "REVISAR"
    gmf = MatchResult(
        transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
        invoice=None,
        status="GMF",
        documento_soporte="SI",
    )
    assert validation_flag(gmf) == "OK"
    assert validation_flag(
        MatchResult(
            transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
            invoice=None,
            status="UNMATCHED",
            documento_soporte="NO",
        )
    ) == "NO"
    assert validation_flag(
        MatchResult(
            transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
            invoice=None,
            status="UNMATCHED",
            documento_soporte="NO",
            suggested_invoice=_rest_demo_invoice(),
        )
    ) == "REVISAR"
    assert validation_flag(
        MatchResult(
            transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
            invoice=None,
            status="AMBIGUOUS",
            documento_soporte="NO",
        )
    ) == "REVISAR"
    assert validation_flag(
        MatchResult(
            transaction=Transaction("1111", date(2026, 5, 1), "X", "COP", 1.0),
            invoice=None,
            status="AMBIGUOUS",
            documento_soporte="NO",
            suggested_invoice=_rest_demo_invoice(),
        )
    ) == "REVISAR"


def test_format_ok_observation_simple():
    """Verifica format ok observation simple."""
    invoice = InvoiceData(
        source_filename="proveedor_cloud.pdf",
        numero_factura="INV-123",
        razon_social="PROVEEDOR CLOUD SAS",
        legible=True,
    )
    match = MatchResult(
        transaction=Transaction("1111", date(2026, 5, 1), "AWS", "COP", 1.0),
        invoice=invoice,
        status="OK",
        documento_soporte="SI",
    )
    assert format_ok_observation(match) == "Factura INV-123 — PROVEEDOR CLOUD SAS"


def test_format_ok_observation_compound():
    """Verifica format ok observation compound."""
    factura = InvoiceData(
        source_filename="f.pdf",
        numero_factura="PS9574533",
        razon_social="RESTAURANTE GAMMA",
        legible=True,
    )
    recibo = InvoiceData(source_filename="r.pdf", numero_factura="805", legible=True)
    match = MatchResult(
        transaction=Transaction("2222", date(2026, 5, 13), "DELIVERY APP", "COP", 1.0),
        invoice=factura,
        status="OK",
        documento_soporte="SI",
        secondary_invoice=recibo,
        match_kind="compound",
    )
    text = format_ok_observation(match)
    assert "PS9574533" in text
    assert "805" in text
    assert "propina" in text


def test_format_ok_observation_multi_factura():
    """Verifica format ok observation multi factura."""
    factura_a = InvoiceData(
        source_filename="REST-CADENA - FE-9821.pdf",
        numero_factura="FE-9821",
        razon_social="RESTAURANTE CADENA S.A.S",
        legible=True,
    )
    factura_b = InvoiceData(
        source_filename="REST-CADENA - FE-9822.pdf",
        numero_factura="FE-9822",
        razon_social="RESTAURANTE CADENA S.A.S",
        legible=True,
    )
    match = MatchResult(
        transaction=Transaction(
            "2222", date(2026, 5, 12), "RESTAURANTE CADENA", "COP", 1.0
        ),
        invoice=factura_a,
        status="OK",
        documento_soporte="NO",
        match_kind="multi_factura",
        component_invoices=[factura_a, factura_b],
    )
    text = format_ok_observation(match)
    assert "FE-9821" in text
    assert "FE-9822" in text
    assert "RESTAURANTE CADENA S.A.S" in text


def test_format_ok_observation_gmf():
    """Verifica format ok observation gmf."""
    match = MatchResult(
        transaction=Transaction(
            "1111",
            date(2026, 5, 11),
            "4X1000 PERSONA JURIDICA",
            "COP",
            225.99,
            is_gmf=True,
        ),
        invoice=None,
        status="GMF",
        documento_soporte="SI",
    )
    text = format_ok_observation(match)
    assert "GMF" in text
    assert "sin factura" in text


def test_format_ok_observation_consolidated_review():
    """Verifica format ok observation consolidated review."""
    from legalizacion_tc.config import Settings

    settings = Settings(
        gcp_project_id="test",
        service_account_email="test@test.iam.gserviceaccount.com",
        plantilla_drive_file_id="",
        control_sheet_id="",
        control_sheet_tab_tarjetas="Tarjetas",
        control_sheet_tab_historico="historico_proveedores",
        historico_drive_folder_id="",
        frankfurter_base_url="https://api.frankfurter.dev/v2",
        amount_tolerance_pct=0.02,
        amount_tolerance_pct_sol=0.12,
        date_tolerance_days=3,
        consolidated_receipt_max_days_after=30,
        consolidated_receipt_review_max_months=3,
        consolidated_max_group_size=6,
        iva_rate_cop=0.19,
        restaurant_no_iva_keywords=(
            "RESTAURANTE",
            "ALMUERZO",
            "CENA",
            "COMIDA",
            "GASTO DE REPRESENTACION",
        ),
        timezone="America/Bogota",
    )
    invoice = InvoiceData(
        source_filename="uber_793.pdf",
        numero_factura="793",
        razon_social="TRANSPORTE APP SAS",
        moneda="COP",
        valor_total_documento=52508.0,
        fecha_factura=date(2026, 5, 20),
        tipo_documento="recibo_caja_menor",
        consolidado=True,
        legible=True,
    )
    match = MatchResult(
        transaction=Transaction(
            "4444",
            date(2026, 3, 31),
            "TRANSPORTE APP",
            "COP",
            52508.0,
        ),
        invoice=invoice,
        status="OK",
        documento_soporte="SI",
        match_kind="consolidated_review",
        needs_review=True,
    )
    text = format_ok_observation(match, settings=settings)
    assert "RECIBO DE CAJA 793" in text
    assert "REVISAR" in text
    assert "50 días" in text


def test_format_ok_observation_provider_date_review(settings):
    """Verifica format ok observation provider date review."""
    invoice = InvoiceData(
        source_filename="restaurante.pdf",
        numero_factura="F-100",
        razon_social="TAKAMI S.A.",
        nombre_comercial="OSAKI ARTISAN",
        moneda="COP",
        valor_total_documento=373120.0,
        fecha_factura=date(2026, 5, 18),
        legible=True,
    )
    match = MatchResult(
        transaction=Transaction(
            "2222",
            date(2026, 6, 18),
            "OSAKI ARTISAN",
            "COP",
            373120.0,
        ),
        invoice=invoice,
        status="OK",
        documento_soporte="SI",
        match_kind="provider_date_review",
        needs_review=True,
    )
    text = format_ok_observation(match, settings=settings)
    assert "F-100" in text
    assert "REVISAR" in text
    assert "31 días" in text
    assert f"±{settings.date_tolerance_days} días" in text


def test_find_near_miss_candidate_sol(settings):
    """Verifica find near miss candidate sol."""
    tx = Transaction(
        card="5555",
        tx_date=date(2026, 5, 24),
        description="RESTAURANTE DEMO",
        currency="COP",
        amount_cop=180000.0,
        amount_original=150.0,
        original_currency="SOL",
        row_index=2,
    )
    candidate = find_near_miss_candidate(settings, tx, [_rest_demo_invoice()])
    assert candidate is not None
    assert candidate.numero_factura == "B050-DEMO001"


def test_reconcile_unmatched_observacion_includes_candidate(settings):
    """Verifica reconcile unmatched observacion includes candidate."""
    extract = ExtractData(
        card="5555",
        period_month="MAYO",
        period_year=2026,
        transactions=[
            Transaction(
                card="5555",
                tx_date=date(2026, 5, 24),
                description="RESTAURANTE DEMO",
                currency="COP",
                amount_cop=180000.0,
                amount_original=150.0,
                original_currency="SOL",
                row_index=2,
            )
        ],
        total_cop=180000.0,
        source_filename="mov.xlsx",
    )
    matches = reconcile(settings, extract, [_rest_demo_invoice()])
    assert matches[0].status == "UNMATCHED"
    assert "B050-DEMO001" in matches[0].observacion
    assert "tolerancia" in matches[0].observacion.lower()
    assert matches[0].failure_reason == "Diferencia de monto"


def test_reconcile_unmatched_sets_suggested_invoice(settings):
    """Verifica reconcile unmatched sets suggested invoice."""
    extract = ExtractData(
        card="5555",
        period_month="MAYO",
        period_year=2026,
        transactions=[
            Transaction(
                card="5555",
                tx_date=date(2026, 5, 24),
                description="RESTAURANTE DEMO",
                currency="COP",
                amount_cop=180000.0,
                amount_original=150.0,
                original_currency="SOL",
                row_index=2,
            )
        ],
        total_cop=180000.0,
        source_filename="mov.xlsx",
    )
    matches = reconcile(settings, extract, [_rest_demo_invoice()])
    assert matches[0].status == "UNMATCHED"
    assert matches[0].suggested_invoice is not None
    assert matches[0].suggested_invoice.numero_factura == "B050-DEMO001"


def test_resolve_suggested_invoice_near_miss(settings):
    """Verifica resolve suggested invoice near miss."""
    tx = Transaction(
        card="5555",
        tx_date=date(2026, 5, 24),
        description="RESTAURANTE DEMO",
        currency="COP",
        amount_cop=180000.0,
        amount_original=150.0,
        original_currency="SOL",
        row_index=2,
    )
    match = MatchResult(
        transaction=tx,
        invoice=None,
        status="UNMATCHED",
        documento_soporte="NO",
    )
    suggested = resolve_suggested_invoice(settings, match, [_rest_demo_invoice()], set())
    assert suggested is not None
    assert suggested.numero_factura == "B050-DEMO001"


def test_pick_best_from_candidates_ambiguous(settings):
    """Verifica pick best from candidates ambiguous."""
    tx = Transaction(
        card="1111",
        tx_date=date(2026, 5, 10),
        description="SERVICIO CLOUD DEMO",
        currency="COP",
        amount_cop=100.0,
        row_index=2,
    )
    inv_a = InvoiceData(
        source_filename="a.pdf",
        numero_factura="A-1",
        razon_social="PROVEEDOR CLOUD SAS",
        moneda="USD",
        valor_total_documento=100.0,
        fecha_factura=date(2026, 5, 10),
        legible=True,
    )
    inv_b = InvoiceData(
        source_filename="b.pdf",
        numero_factura="B-1",
        razon_social="Other Corp",
        moneda="USD",
        valor_total_documento=100.0,
        fecha_factura=date(2026, 5, 10),
        legible=True,
    )
    best = pick_best_from_candidates(settings, tx, [inv_b, inv_a])
    assert best is not None
    assert best.numero_factura == "A-1"


def test_find_amount_date_suggestion_candidate(settings):
    """Verifica find amount date suggestion candidate."""
    tx = Transaction(
        card="5555",
        tx_date=date(2026, 5, 8),
        description="INVERSIONES ALPHA COL",
        currency="COP",
        amount_cop=75280.0,
        row_index=2,
    )
    culto = InvoiceData(
        source_filename="CAFE BETA S.A.S - CBB29786.pdf",
        numero_factura="CBB29786",
        razon_social="CAFE BETA S.A.S",
        moneda="COP",
        valor_total_documento=75279.63,
        fecha_factura=date(2026, 5, 8),
        legible=True,
    )
    ktronix = InvoiceData(
        source_filename="KTRONIX.pdf",
        numero_factura="X8472527982",
        razon_social="Colombiana de Comercio S.A.",
        moneda="COP",
        valor_total_documento=347315.0,
        fecha_factura=date(2026, 5, 8),
        legible=True,
    )
    from legalizacion_tc.validation_notes import find_amount_date_suggestion_candidate

    suggested = find_amount_date_suggestion_candidate(
        settings, tx, [culto, ktronix], set()
    )
    assert suggested is not None
    assert suggested.numero_factura == "CBB29786"


def test_find_near_miss_skips_reserved_invoices(settings):
    """Verifica find near miss skips reserved invoices."""
    tx = Transaction(
        card="5555",
        tx_date=date(2026, 5, 24),
        description="REST DEMO",
        currency="COP",
        amount_cop=185000.0,
        amount_original=155.0,
        original_currency="SOL",
        row_index=3,
    )
    sushi = InvoiceData(
        source_filename="RESTAURANTE DEMO -F001.jpeg",
        numero_factura="B050-DEMO001",
        razon_social="RESTAURANTE DEMO S.A.S",
        moneda="SOL",
        valor_total_documento=129.0,
        fecha_factura=date(2026, 5, 24),
        detalle_gasto="TC 5555 GASTO REST DEMO 24 DE MAYO",
        legible=True,
    )
    assert (
        find_near_miss_candidate(settings, tx, [sushi], {sushi.source_filename})
        is None
    )


def test_pick_best_from_candidates_skips_reserved(settings):
    """Verifica pick best from candidates skips reserved."""
    tx = Transaction(
        card="1111",
        tx_date=date(2026, 5, 10),
        description="SERVICIO CLOUD DEMO",
        currency="COP",
        amount_cop=100.0,
        row_index=2,
    )
    inv_a = InvoiceData(
        source_filename="a.pdf",
        numero_factura="A-1",
        razon_social="PROVEEDOR CLOUD SAS",
        moneda="USD",
        valor_total_documento=100.0,
        fecha_factura=date(2026, 5, 10),
        legible=True,
    )
    inv_b = InvoiceData(
        source_filename="b.pdf",
        numero_factura="B-1",
        razon_social="Other Corp",
        moneda="USD",
        valor_total_documento=100.0,
        fecha_factura=date(2026, 5, 10),
        legible=True,
    )
    assert (
        pick_best_from_candidates(settings, tx, [inv_a, inv_b], {inv_a.source_filename})
        is None
    )


def test_reconcile_ok_observacion_includes_invoice_number(settings):
    """Verifica reconcile ok observacion includes invoice number."""
    extract = ExtractData(
        card="5555",
        period_month="MAYO",
        period_year=2026,
        transactions=[
            Transaction(
                card="5555",
                tx_date=date(2026, 5, 24),
                description="RESTAURANTE DEMO",
                currency="COP",
                amount_cop=155091.88,
                amount_original=141.9,
                original_currency="SOL",
                row_index=2,
            )
        ],
        total_cop=155091.88,
        source_filename="mov.xlsx",
    )
    matches = reconcile(settings, extract, [_rest_demo_invoice()])
    assert matches[0].status == "OK"
    assert "B050-DEMO001" in matches[0].observacion


@pytest.fixture
def sample_extract_xlsx(tmp_path):
    """Fixture o helper: sample extract xlsx."""
    path = tmp_path / "mov.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Concepto", "Moneda", "Monto", VALIDATION_COLUMN])
    ws.append(["2026-05-24", "RESTAURANTE DEMO S.A.S", "SOL", 141.9, ""])
    wb.save(path)
    wb.close()
    return path


def test_apply_extract_review_columns_writes_ok_no_and_observations(
    sample_extract_xlsx,
):
    """Verifica apply extract review columns writes ok no and observations."""
    matches = [
        MatchResult(
            transaction=Transaction(
                card="5555",
                tx_date=date(2026, 5, 24),
                description="RESTAURANTE DEMO",
                currency="COP",
                amount_cop=155091.88,
                row_index=2,
            ),
            invoice=_rest_demo_invoice(),
            status="UNMATCHED",
            documento_soporte="NO",
            observacion="Sin match: monto 9.0% — candidato B050-DEMO001",
        ),
    ]
    apply_extract_review_columns(sample_extract_xlsx, matches)
    wb = load_workbook(sample_extract_xlsx)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert OBSERVATIONS_COLUMN in headers
    assert ws.cell(2, headers.index(VALIDATION_COLUMN) + 1).value == "NO"
    assert "B050-DEMO001" in str(
        ws.cell(2, headers.index(OBSERVATIONS_COLUMN) + 1).value
    )
    wb.close()
